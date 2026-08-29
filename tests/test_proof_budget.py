"""The proof budget must bound work, never bound what the tool claims to have seen.

An earlier version capped by slicing the findings list. That deleted detections:
on four corpus workbooks the deleted findings included the injected errors, so a
runtime cap was scored as a detector failure. These tests pin the invariant.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from plumbline.audit import audit

ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture(scope="module")
def two_breaks(tmp_path_factory) -> Path:
    """A sheet with two independent row-pattern breaks, so a cap of 1 bites."""
    path = tmp_path_factory.mktemp("budget") / "two_breaks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(2, 8):
        for col, letter in enumerate("BCDE", start=2):
            ws.cell(row=row, column=col, value=row * col)
        # F..I total the four columns above; two rows total only three.
        for col in range(6, 10):
            short = row in (4, 6) and col == 6
            top, bottom = 2, (3 if short else 4)
            ws.cell(row=row, column=col, value=f"=SUM({chr(64 + col - 4)}{top}:{chr(64 + col - 4)}{bottom})")
    wb.save(path)
    return path


@pytest.fixture(scope="module")
def uncapped(two_breaks):
    return audit(two_breaks, check_determinism=False)


@pytest.fixture(scope="module")
def capped(two_breaks):
    return audit(two_breaks, check_determinism=False, max_proofs=1)


class TestBudgetPreservesDetection:
    def test_same_cells_are_reported_either_way(self, uncapped, capped):
        assert {(f.sheet, f.cell) for f in capped.findings} == {
            (f.sheet, f.cell) for f in uncapped.findings
        }

    def test_detected_count_is_unaffected(self, uncapped, capped):
        assert capped.detected == uncapped.detected
        assert len(capped.findings) == len(uncapped.findings)


class TestBudgetIsVisible:
    def test_a_capped_audit_says_so(self, capped, uncapped):
        # The fixture has more than one finding, else the cap is untested.
        assert len(uncapped.findings) > 1
        assert capped.proof_truncated is True
        assert capped.proof_deferred == len(uncapped.findings) - 1

    def test_deferred_findings_are_unproved_and_explain_why(self, capped):
        deferred = [f for f in capped.findings if "budget" in f.proof]
        assert deferred, "expected at least one deferred finding"
        assert all(not f.proved for f in deferred)

    def test_an_uncapped_audit_reports_no_deferral(self, uncapped):
        assert uncapped.proof_truncated is False
        assert uncapped.proof_deferred == 0

    def test_serialised_report_carries_the_deferral(self, capped):
        d = capped.to_dict()
        assert d["proof_truncated"] is True
        assert d["proof_deferred"] >= 1


class TestBudgetDoesNotFabricateProofs:
    def test_deferral_is_never_mistaken_for_a_clean_result(self):
        """A workbook with nothing wrong and a workbook that ran out of budget
        must not serialise the same way."""
        wb = Workbook()
        ws = wb.active
        for r in range(1, 5):
            ws.cell(row=r, column=1, value=r)
            ws.cell(row=r, column=2, value=f"=A{r}*2")
        path = Path(__file__).parent / "fixtures" / "_budget_clean.xlsx"
        wb.save(path)
        try:
            clean = audit(path, check_determinism=False)
            assert clean.proof_truncated is False
            assert clean.proof_deferred == 0
        finally:
            path.unlink(missing_ok=True)
