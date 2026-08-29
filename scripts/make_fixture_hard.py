"""Build the deliberately hard fixture: a subtotal that is right today and wrong tomorrow.

Practitioner literature says the errors that survive review are the ones that
"hide in places that feel like they don't need checking -- subtotals that balance
despite incorrect line items beneath them."

This fixture encodes the nastiest realistic version of that. Q2's Total Opex is a
typed-in constant (30000) rather than a formula. Today it exactly equals the sum
of the lines beneath it, so:

  * every value-based check passes -- the total is correct
  * the arithmetic ties -- Operating Income is right
  * a human scanning the numbers sees nothing wrong

But the cell is dead. Change any Opex input and the total silently stops tracking.
That is Panko's "hardcoding" class, and it is the error that detonates later, in
someone else's hands, after the model has been signed off.

Detecting it needs a *different proof technique* from the off-by-one fixture: you
cannot show a delta by repairing the formula, because the current value is already
correct. You have to perturb an upstream input and show the dependent fails to
respond. See `sensitivity_probe` in the detector suite.

Usage:  python scripts/make_fixture_hard.py
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

FIXTURE_DIR = Path(__file__).resolve().parents[1] / "tests" / "fixtures"
QUARTERS = ["B", "C", "D", "E"]
SEEDED_COLUMN = "C"

OPEX_ROWS = {
    8: ("Salaries", [20000, 21000, 22000, 23000]),
    9: ("Marketing", [5000, 6000, 7000, 8000]),
    10: ("Rent", [3000, 3000, 3000, 3000]),
}


def build() -> tuple[Path, Path]:
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    ws["A1"] = "Quarterly P&L"
    for col, label in zip(QUARTERS, ["Q1", "Q2", "Q3", "Q4"]):
        ws[f"{col}2"] = label

    for row, (label, values) in {
        3: ("Revenue", [100000, 120000, 135000, 150000]),
        4: ("COGS", [40000, 48000, 54000, 60000]),
        **OPEX_ROWS,
    }.items():
        ws[f"A{row}"] = label
        for col, value in zip(QUARTERS, values):
            ws[f"{col}{row}"] = value

    ws["A5"] = "Gross Profit"
    ws["A7"] = "Operating Expenses"
    ws["A11"] = "Total Opex"
    ws["A13"] = "Operating Income"

    # The value the hardcode must match to stay invisible: 21000 + 6000 + 3000.
    seeded_index = QUARTERS.index(SEEDED_COLUMN)
    correct_total = sum(values[seeded_index] for _, values in OPEX_ROWS.values())

    for col in QUARTERS:
        ws[f"{col}5"] = f"={col}3-{col}4"
        if col == SEEDED_COLUMN:
            ws[f"{col}11"] = correct_total  # <-- dead constant, correct today
        else:
            ws[f"{col}11"] = f"=SUM({col}8:{col}10)"
        ws[f"{col}13"] = f"={col}5-{col}11"

    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = FIXTURE_DIR / "quarterly_pl_hardcoded.xlsx"
    wb.save(xlsx_path)

    manifest = {
        "workbook": xlsx_path.name,
        "difficulty": "hard",
        "why_hard": (
            "The hardcoded total equals the correct sum today, so every value-based "
            "check passes and the arithmetic ties. Only a structural check (constant "
            "where the row expects a formula) or a sensitivity probe (perturb an input, "
            "observe the total fail to respond) can catch it."
        ),
        "seeded_errors": [
            {
                "id": "pl-opex-hardcoded",
                "sheet": "P&L",
                "cell": f"{SEEDED_COLUMN}11",
                "panko_class": "hardcoding",
                "description": (
                    "Total Opex is a typed constant, not a SUM. It matches the lines "
                    "beneath it today and silently diverges the moment any input changes."
                ),
                "actual_value": correct_total,
                "expected_formula": f"=SUM({SEEDED_COLUMN}8:{SEEDED_COLUMN}10)",
                "detectable_by": ["pattern_break", "sensitivity_probe"],
                "detectable_by_value_check": False,
                "propagates_to": [f"{SEEDED_COLUMN}13"],
            }
        ],
    }
    manifest_path = FIXTURE_DIR / "quarterly_pl_hardcoded.truth.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return xlsx_path, manifest_path


if __name__ == "__main__":
    import argparse
    import sys

    # These scripts overwrite committed test fixtures, so they must never run as a
    # side effect of someone poking at them -- `make_fixture.py --help` silently
    # rewrote both fixtures once. Regenerating is now something you ask for.
    ap = argparse.ArgumentParser(
        description="Regenerate a committed test fixture. Overwrites files under "
                    "tests/fixtures/; pass --write to confirm."
    )
    ap.add_argument("--write", action="store_true", help="actually overwrite the fixture")
    args = ap.parse_args()
    if not args.write:
        print("refusing to overwrite committed fixtures without --write", file=sys.stderr)
        raise SystemExit(2)

    xlsx, manifest = build()
    print(f"wrote {xlsx}")
    print(f"wrote {manifest}")
