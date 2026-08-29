"""Build a small fixture workbook carrying one deliberately seeded error.

The error is Panko's "pointing" class (mechanical): an off-by-one SUM range that
silently omits the Rent line from Q2's Total Opex. Every other quarter sums the
full block, so the mistake is a *pattern break* -- detectable structurally -- and
it propagates into Operating Income, so it is materially wrong, not cosmetic.

Usage:  python scripts/make_fixture.py
Writes: tests/fixtures/quarterly_pl.xlsx  and  a ground-truth manifest beside it.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

FIXTURE_DIR = Path(__file__).resolve().parents[1] / "tests" / "fixtures"
QUARTERS = ["B", "C", "D", "E"]

# The seeded fault: Q2 (column C) sums only rows 8:9, omitting Rent on row 10.
SEEDED_COLUMN = "C"
CORRECT_OPEX_RANGE = "{col}8:{col}10"
SEEDED_OPEX_RANGE = "{col}8:{col}9"


def build() -> tuple[Path, Path]:
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    ws["A1"] = "Quarterly P&L"
    for col, label in zip(QUARTERS, ["Q1", "Q2", "Q3", "Q4"]):
        ws[f"{col}2"] = label

    rows = {
        3: ("Revenue", [100000, 120000, 135000, 150000]),
        4: ("COGS", [40000, 48000, 54000, 60000]),
        8: ("Salaries", [20000, 21000, 22000, 23000]),
        9: ("Marketing", [5000, 6000, 7000, 8000]),
        10: ("Rent", [3000, 3000, 3000, 3000]),
    }
    for row, (label, values) in rows.items():
        ws[f"A{row}"] = label
        for col, value in zip(QUARTERS, values):
            ws[f"{col}{row}"] = value

    ws["A5"] = "Gross Profit"
    ws["A7"] = "Operating Expenses"
    ws["A11"] = "Total Opex"
    ws["A13"] = "Operating Income"

    for col in QUARTERS:
        ws[f"{col}5"] = f"={col}3-{col}4"
        opex = SEEDED_OPEX_RANGE if col == SEEDED_COLUMN else CORRECT_OPEX_RANGE
        ws[f"{col}11"] = f"=SUM({opex.format(col=col)})"
        ws[f"{col}13"] = f"={col}5-{col}11"

    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = FIXTURE_DIR / "quarterly_pl.xlsx"
    wb.save(xlsx_path)

    # Ground truth, so the evaluation never depends on remembering what we seeded.
    manifest = {
        "workbook": xlsx_path.name,
        "seeded_errors": [
            {
                "id": "pl-opex-offbyone",
                "sheet": "P&L",
                "cell": f"{SEEDED_COLUMN}11",
                "panko_class": "mechanical/pointing",
                "description": "SUM range omits the Rent line (row 10) that every other quarter includes.",
                "actual_formula": f"=SUM({SEEDED_OPEX_RANGE.format(col=SEEDED_COLUMN)})",
                "expected_formula": f"=SUM({CORRECT_OPEX_RANGE.format(col=SEEDED_COLUMN)})",
                "propagates_to": [f"{SEEDED_COLUMN}13"],
            }
        ],
    }
    manifest_path = FIXTURE_DIR / "quarterly_pl.truth.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return xlsx_path, manifest_path


if __name__ == "__main__":
    import argparse
    import sys

    # These scripts overwrite committed test fixtures, so they must never run as a
    # side effect of someone poking at them -- `make_fixture.py --help` silently
    # rewrote both fixtures once. Regenerating is now something you ask for.
    ap = argparse.ArgumentParser(
        description="Regenerate a committed test fixture. Overwrites files under "
                    "tests/fixtures/; pass --write to confirm."
    )
    ap.add_argument("--write", action="store_true", help="actually overwrite the fixture")
    args = ap.parse_args()
    if not args.write:
        print("refusing to overwrite committed fixtures without --write", file=sys.stderr)
        raise SystemExit(2)

    xlsx, manifest = build()
    print(f"wrote {xlsx}")
    print(f"wrote {manifest}")
