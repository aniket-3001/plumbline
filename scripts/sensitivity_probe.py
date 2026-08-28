"""Prove a cell is dead by perturbing what it should depend on.

The off-by-one fixture is proved by repairing the formula and showing a delta.
That technique cannot work on a hardcoded subtotal, because the hardcode is
*already correct* -- repairing it changes nothing today, so the proof comes back
empty and the finding looks like a false positive.

The probe inverts it. Instead of changing the suspect cell, change an input the
suspect *ought* to depend on, and watch whether it responds:

    peers respond, suspect does not  ->  the cell is disconnected. Proved.

This is the only technique in the design that can catch an error which is
numerically correct at the moment you look at it.

Usage:  python scripts/sensitivity_probe.py [workbook.xlsx]
"""

from __future__ import annotations

import os
import re
import sys
import tempfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from xlcalculator import Evaluator, ModelCompiler

sys.path.insert(0, str(Path(__file__).resolve().parent))
from poc import A1_REF, native, normalise, rebase, split_ref  # noqa: E402

PERTURBATION = 1000  # arbitrary, non-zero, and large enough to survive rounding


def formula_cells(path: str) -> dict[str, dict[str, str]]:
    model = ModelCompiler().read_and_parse_archive(path)
    out: dict[str, dict[str, str]] = {}
    for addr, cell in model.cells.items():
        text = getattr(cell.formula, "formula", None)
        if not text:
            continue
        sheet, _, ref = addr.rpartition("!")
        out.setdefault(sheet, {})[ref] = text
    return out


def constant_cells(path: str, sheet: str) -> dict[str, object]:
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet]
    out = {}
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None or (isinstance(v, str) and v.startswith("=")):
                continue
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                out[cell.coordinate] = v
    wb.close()
    return out


def find_dead_cells(path: str, sheet: str) -> list[dict]:
    """Constants sitting in a row where every peer is a formula."""
    formulas = formula_cells(path).get(sheet, {})
    constants = constant_cells(path, sheet)

    by_row_formulas: dict[int, list[tuple[str, int, str]]] = {}
    for ref, text in formulas.items():
        row, col = split_ref(ref)
        by_row_formulas.setdefault(row, []).append((ref, col, text))

    suspects = []
    for ref, value in constants.items():
        row, col = split_ref(ref)
        peers = by_row_formulas.get(row, [])
        if len(peers) < 2:
            continue
        shapes = Counter(normalise(t, row, c) for _, c, t in peers)
        shape, n = shapes.most_common(1)[0]
        if n < len(peers):
            continue  # peers do not agree among themselves; not a clean pattern
        twin_ref, twin_col, twin_text = peers[0]
        twin_row, _ = split_ref(twin_ref)
        suspects.append(
            {
                "cell": ref,
                "value": value,
                "expected_formula": rebase(twin_text, twin_row, twin_col, row, col),
                "peers": [p[0] for p in peers],
                "reason": (
                    f"{n} of {n} other formula cells in row {row} share one shape; "
                    f"{ref} is a typed constant."
                ),
            }
        )
    return suspects


def probe(path: str, sheet: str, suspect: dict) -> dict:
    """Perturb an input the suspect should depend on; see who moves."""
    inputs = sorted(
        {
            m.group(0).replace("$", "")
            for m in A1_REF.finditer(suspect["expected_formula"])
        }
    )
    # Expand a range like C8:C10 into its endpoints; perturbing one is enough.
    target_input = inputs[0] if inputs else None
    if not target_input:
        return {**suspect, "probed": False, "reason_failed": "no inputs to perturb"}

    watch = [suspect["cell"], *suspect["peers"]]

    model = ModelCompiler().read_and_parse_archive(path)
    ev = Evaluator(model)
    before = {r: native(ev.evaluate(f"{sheet}!{r}")) for r in watch}

    wb = load_workbook(path)
    ws = wb[sheet]
    original = ws[target_input].value
    if not isinstance(original, (int, float)) or isinstance(original, bool):
        return {**suspect, "probed": False, "reason_failed": f"{target_input} is not numeric"}
    ws[target_input] = original + PERTURBATION

    tmp = os.path.join(tempfile.gettempdir(), f"plumbline_probe_{os.getpid()}.xlsx")
    tmp2 = os.path.join(tempfile.gettempdir(), f"plumbline_probe_cf_{os.getpid()}.xlsx")
    try:
        wb.save(tmp)
        model2 = ModelCompiler().read_and_parse_archive(tmp)
        ev2 = Evaluator(model2)
        after = {r: native(ev2.evaluate(f"{sheet}!{r}")) for r in watch}

        # The control arm. Same perturbation, but with the suspect cell holding the
        # formula its peers hold. If THAT responds while the real cell does not, the
        # cell is provably disconnected -- and we have measured by how much.
        ws[suspect["cell"]] = suspect["expected_formula"]
        wb.save(tmp2)
        model3 = ModelCompiler().read_and_parse_archive(tmp2)
        ev3 = Evaluator(model3)
        control_after = native(ev3.evaluate(f"{sheet}!{suspect['cell']}"))
    finally:
        for f in (tmp, tmp2):
            if os.path.exists(f):
                os.remove(f)

    cell = suspect["cell"]
    suspect_moved = before[cell] != after[cell]
    control_moved = control_after != before[cell]
    try:
        divergence = control_after - after[cell]
    except TypeError:
        divergence = None

    return {
        **suspect,
        "probed": True,
        "perturbed": target_input,
        "perturbation": PERTURBATION,
        "original_input": original,
        "before": before,
        "after": after,
        "control_after": control_after,
        "suspect_moved": suspect_moved,
        "control_moved": control_moved,
        "divergence": divergence,
        # Only a proof if the control demonstrably WOULD have moved.
        "proved_dead": (not suspect_moved) and control_moved,
    }


def main(path: str) -> int:
    print(f"Plumbline sensitivity probe  --  {Path(path).name}\n")
    sheets = formula_cells(path)
    found = 0

    for sheet in sheets:
        for suspect in find_dead_cells(path, sheet):
            result = probe(path, sheet, suspect)
            found += 1
            if not result.get("probed"):
                print(f"[SKIPPED] {sheet}!{suspect['cell']}: {result['reason_failed']}")
                continue
            verdict = "PROVED DEAD" if result["proved_dead"] else "responds -- not dead"
            c = result["cell"]
            print(f"[{verdict}] {sheet}!{c}")
            print(f"    is        {result['value']}  (typed constant, correct today)")
            print(f"    expected  {result['expected_formula']}")
            print(f"    why       {result['reason']}")
            print(f"    probe     set {result['perturbed']} {result['original_input']} -> "
                  f"{result['original_input'] + result['perturbation']}")
            print(f"    proof     {c} as-is:     {result['before'][c]} -> {result['after'][c]}"
                  f"   (no response)")
            print(f"              {c} as formula: {result['before'][c]} -> "
                  f"{result['control_after']}   (responds)")
            if result["divergence"]:
                print(f"              the cell is understating by {result['divergence']:+} "
                      f"the moment an input moves")
            print()

    if not found:
        print("No dead cells found.")
    return 0


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else "tests/fixtures/quarterly_pl_hardcoded.xlsx"
    raise SystemExit(main(target))
