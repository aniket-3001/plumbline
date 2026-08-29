"""The direct-prompt LLM baseline: hand the model the sheet and ask it to find errors.

    python scripts/run_llm_baseline.py --limit 5          # cost-controlled trial
    python scripts/run_llm_baseline.py                    # full corpus

This is the brief's named baseline shape -- "one direct prompt with basic
instructions" -- and it is the arm that answers the question the deterministic
numbers cannot: **does the structural machinery earn its place, or would asking a
good model directly have worked just as well?**

Three commitments make the comparison worth reading.

**It is not a strawman.** Same model this project uses for Layer 2, at the same
default effort. Same workbooks, same seeds, same scorer, same difficulty
stratification. The model is given the formulas in reading order with row labels,
which is the information a human auditor would have. A weak baseline here would
flatter the deterministic arm and teach us nothing.

**It gets its own exclusion list, computed the same way.** Enron's files are full of
pre-existing anomalies with no ground truth, and the deterministic arms are not
charged for the ones they find. Neither is this one -- but "what this arm considers
pre-existing" is a property of *this* arm, so the model is run twice: once on the
untouched original and once on the seeded copy. Anything it flags on the original is
excluded. Anything else would charge the model for noticing real Enron oddities,
which is exactly the bug that cost the deterministic arm 11 false positives.

**Its failures are recorded, not summarised.** Every raw reply is written to
`results/llm_baseline_raw/`, so a claim about what the model did can be checked
against what it actually said.

Cost is printed as it goes and the run can be capped. A full pass is roughly 900K
input tokens across both halves.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "scripts"))

SEEDED = ROOT / "data" / "seeded"
EVAL = ROOT / "data" / "eval_corpus"
RESULTS = ROOT / "results"
RAW = RESULTS / "llm_baseline_raw"

#: Formulas per request. Large enough that the model sees a whole block of the sheet
#: and can spot a break in it; small enough that a 10,000-formula workbook does not
#: become one enormous request whose middle the model skims.
CHUNK = 400

#: Opus 5, $ per million tokens.
IN_PER_M, OUT_PER_M = 5.00, 25.00

PROMPT = """\
You are auditing a spreadsheet for errors. Below are the formulas from one sheet, \
in reading order, with the row label where one exists.

Find cells that are wrong. The errors that matter most are:
- a formula that breaks the pattern of the cells around it (an off-by-one range, a \
reference pointing one row or column away from where its neighbours point)
- a typed-in constant sitting where the surrounding cells hold formulas, so the \
value no longer updates when its inputs change
- a formula that computes something other than what its label says

Report only cells you believe are actually wrong. A sheet may contain no errors at \
all, and reporting a correct cell is worse than missing one.

Reply as JSON only, with no other text:
{"findings": [{"cell": "<address, e.g. C11>", "reason": "<one sentence>"}]}"""

CELL = re.compile(r"^[A-Z]{1,3}[1-9][0-9]{0,6}$")


def sheet_chunks(path: Path):
    """(sheet, text) blocks, each a readable slice of one sheet."""
    from openpyxl import load_workbook

    from poc import load_formulas

    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        labels: dict[str, dict[int, str]] = {}
        for name in wb.sheetnames:
            ws = wb[name]
            per_row: dict[int, str] = {}
            for row in ws.iter_rows(max_col=3):
                for cell in row:
                    if isinstance(cell.value, str) and not cell.value.startswith("="):
                        per_row.setdefault(cell.row, cell.value[:40])
            labels[name] = per_row
    finally:
        wb.close()

    for sheet, formulas in load_formulas(str(path)).items():
        def sort_key(ref: str):
            m = re.match(r"([A-Z]+)([0-9]+)", ref)
            return (int(m.group(2)), m.group(1)) if m else (0, ref)

        refs = sorted(formulas, key=sort_key)
        for i in range(0, len(refs), CHUNK):
            block = refs[i : i + CHUNK]
            lines, seen = [], None
            for ref in block:
                row = int(re.sub(r"[^0-9]", "", ref) or 0)
                label = labels.get(sheet, {}).get(row)
                if label and row != seen:
                    lines.append(f"  [row {row}: {label}]")
                    seen = row
                lines.append(f"  {ref}: {formulas[ref]}")
            yield sheet, "\n".join(lines)


def audit_with_model(path: Path, call, tag: str, budget: dict) -> set[tuple[str, str]]:
    """Every (sheet, cell) the model calls wrong. Raw replies are kept."""
    found: set[tuple[str, str]] = set()
    RAW.mkdir(parents=True, exist_ok=True)

    for n, (sheet, text) in enumerate(sheet_chunks(path), 1):
        user = f"Sheet: {sheet}\n\nFormulas:\n{text}"
        try:
            reply, usage = call(PROMPT, user)
        except Exception as exc:  # noqa: BLE001 -- one bad chunk must not kill the run
            message = str(exc)
            # ...but some failures are not per-chunk, and retrying them is worse than
            # useless. When the account runs out of credit every remaining request
            # fails identically; the first run of this script logged 76 such failures
            # before it was killed by hand. Fatal conditions abort immediately.
            fatal = (
                "credit balance is too low" in message
                or "authentication_error" in message
                or "invalid x-api-key" in message
                or "permission_error" in message
            )
            print(f"      chunk {n} failed: {type(exc).__name__}: {message[:120]}", flush=True)
            if fatal:
                raise RuntimeError(f"aborting: {message[:200]}") from exc
            continue

        budget["in"] += usage[0]
        budget["out"] += usage[1]
        stem = f"{path.stem[:40]}__{tag}__{n:03d}".replace(" ", "_")
        (RAW / f"{stem}.json").write_text(
            json.dumps({"sheet": sheet, "prompt": user, "reply": reply}, indent=2),
            encoding="utf-8", newline="\n",
        )

        body = reply.strip()
        if body.startswith("```"):
            body = re.sub(r"^```[a-zA-Z]*\n", "", body)
            body = re.sub(r"\n```$", "", body).strip()
        try:
            payload = json.loads(body)
        except json.JSONDecodeError:
            print(f"      chunk {n}: unparsable reply, skipped", flush=True)
            continue

        for item in payload.get("findings") or []:
            ref = str(item.get("cell", "")).upper().replace("$", "")
            if CELL.match(ref):
                found.add((sheet, ref))
    return found


def make_call(model: str):
    import anthropic

    client = anthropic.Anthropic()

    def call(system: str, user: str):
        resp = client.messages.create(
            model=model,
            max_tokens=4000,
            system=system,
            thinking={"type": "adaptive"},
            output_config={"effort": "medium"},
            messages=[{"role": "user", "content": user}],
        )
        text = "" if resp.stop_reason == "refusal" else "".join(
            b.text for b in resp.content if getattr(b, "type", "") == "text"
        )
        return text, (resp.usage.input_tokens, resp.usage.output_tokens)

    return call


def main(argv=None) -> int:
    from plumbline.agent import DEFAULT_MODEL
    from plumbline.scoring import Scorecard, score

    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--limit", type=int, default=0, help="only run N workbooks")
    ap.add_argument(
        "--max-usd",
        type=float,
        default=2.00,
        help="hard spend ceiling; the run stops cleanly when the measured cost "
        "reaches it and writes what it has. Raise it deliberately -- this default "
        "is low because an unattended batch job spends real money, and the first "
        "full run of this script drained an account mid-corpus.",
    )
    ap.add_argument("--model", default=DEFAULT_MODEL)
    ap.add_argument("--out", default="llm_baseline.json")
    ap.add_argument(
        "--skip-originals",
        action="store_true",
        help="do not compute this arm's exclusion list (halves cost; makes the "
        "comparison unfair to this arm, and is recorded as such in the output)",
    )
    args = ap.parse_args(argv)

    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ANTHROPIC_API_KEY is not set", file=sys.stderr)
        return 2

    manifests = sorted(SEEDED.glob("*.truth.json"))
    if not manifests:
        print("no seeded workbooks -- run scripts/seed_corpus.py first", file=sys.stderr)
        return 1
    if args.limit:
        manifests = manifests[: args.limit]

    call = make_call(args.model)
    budget = {"in": 0, "out": 0}
    total, rows = Scorecard(), []
    started = time.time()

    print(f"direct-prompt baseline: {len(manifests)} workbooks, {args.model}\n", flush=True)

    stopped_early = None
    for mpath in manifests:
        manifest = json.loads(mpath.read_text(encoding="utf-8"))
        seeded = SEEDED / manifest["seeded"]
        if not seeded.exists():
            continue

        spent = budget["in"] / 1e6 * IN_PER_M + budget["out"] / 1e6 * OUT_PER_M
        if spent >= args.max_usd:
            stopped_early = (f"spend ceiling reached: ${spent:.2f} of ${args.max_usd:.2f} "
                             f"after {len(rows)} of {len(manifests)} workbooks")
            print(f"\nSTOPPING: {stopped_early}", flush=True)
            print("  Raise --max-usd to continue further.", flush=True)
            break
        t0 = time.time()

        try:
            flagged = audit_with_model(seeded, call, "seeded", budget)
        except RuntimeError as exc:
            stopped_early = str(exc)
            print(f"\n{exc}", flush=True)
            break

        # This arm's own exclusion list: what it flags on the *untouched* original.
        pre: set[tuple[str, str]] = set()
        original = Path(manifest["source"])
        if not original.exists():
            original = EVAL / Path(manifest["workbook"]).name
        if not args.skip_originals and original.exists():
            try:
                pre = audit_with_model(original, call, "original", budget)
            except RuntimeError as exc:
                # The seeded half already cost money; discard this workbook rather
                # than score it against an exclusion list that was never finished.
                stopped_early = str(exc)
                print(f"\n{exc}", flush=True)
                break

        scoped = dict(manifest)
        scoped["pre_existing_findings"] = [f"{s}!{c}" for s, c in pre]
        payload = [{"sheet": s, "cell": c, "proved": False} for s, c in flagged]
        card = score(payload, scoped, require_proof=False)
        total = total.merge(card)

        spend = budget["in"] / 1e6 * IN_PER_M + budget["out"] / 1e6 * OUT_PER_M
        rows.append({
            "workbook": seeded.name,
            "seeds": manifest["seed_count"],
            "reported": len(flagged),
            "pre_existing_by_this_arm": len(pre),
            "seconds": round(time.time() - t0, 1),
            **card.to_dict(),
        })
        print(f"  {seeded.name[:44]:46} reported {len(flagged):3d}  "
              f"tp {card.true_positives:2d}  fp {card.false_positives:3d}  "
              f"pre {len(pre):3d}  ${spend:6.2f}", flush=True)

        RESULTS.mkdir(exist_ok=True)
        (RESULTS / args.out).write_text(
            json.dumps({"summary": total.to_dict(), "workbooks": rows}, indent=2),
            encoding="utf-8", newline="\n",
        )

    spend = budget["in"] / 1e6 * IN_PER_M + budget["out"] / 1e6 * OUT_PER_M
    summary = total.to_dict()
    summary.update({
        "arm": "direct_prompt_llm",
        "model": args.model,
        "workbooks": len(rows),
        "input_tokens": budget["in"],
        "output_tokens": budget["out"],
        "usd": round(spend, 2),
        "exclusions_computed": not args.skip_originals,
        "max_usd": args.max_usd,
        "stopped_early": stopped_early,
        "complete": stopped_early is None and len(rows) == len(manifests),
        "total_seconds": round(time.time() - started, 1),
    })
    (RESULTS / args.out).write_text(
        json.dumps({"summary": summary, "workbooks": rows}, indent=2),
        encoding="utf-8", newline="\n",
    )

    print("\n" + "=" * 66)
    print("  ARM: direct-prompt LLM")
    print("=" * 66)
    print(f"  workbooks           {len(rows)}"
          + ("" if summary["complete"] else f"  ** PARTIAL: {stopped_early or 'run did not finish'} **"))
    print(f"  found               {total.true_positives}")
    print(f"  missed              {total.false_negatives}")
    print(f"  false positives     {total.false_positives}")
    print(f"  excluded            {total.pre_existing_hits}  (flagged on the original too)")
    print()
    print(f"  precision           {total.precision:.3f}")
    print(f"  recall              {total.recall:.3f}")
    print(f"  F1                  {total.f1:.3f}")
    print()
    print("  recall by difficulty:")
    for d in ("obvious", "realistic", "silent"):
        b = total.by_difficulty.get(d, {"found": 0, "missed": 0})
        if b["found"] + b["missed"]:
            print(f"    {d:10} {total.recall_for(d):.3f}   ({b['found']}/{b['found'] + b['missed']})")
    print()
    print(f"  tokens              {budget['in']:,} in / {budget['out']:,} out")
    print(f"  cost                ${spend:.2f}")
    print(f"  wall clock          {summary['total_seconds']}s")
    print(f"\nwrote results/{args.out}; raw replies in results/llm_baseline_raw/")
    return 0


if __name__ == "__main__":
    sys.exit(main())
