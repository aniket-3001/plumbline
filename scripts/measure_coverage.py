"""Measure how much of the Enron corpus Plumbline can actually evaluate.

This answers the question that decides whether the whole design is viable:
when real financial models use functions xlcalculator does not implement, how
much of the corpus falls out of reach?

Two passes:

  1. **Function census** (fast, openpyxl): read formulas without evaluating them,
     extract every function name, and compare against xlcalculator's registry.
     Gives per-workbook and per-formula-cell supportability.

  2. **Parse check** (slow, xlcalculator): on a sample, confirm the model actually
     compiles. Supportable-on-paper is not the same as parses-in-practice.

Usage:
    python scripts/measure_coverage.py --limit 500
    python scripts/measure_coverage.py --limit 500 --parse-sample 50
"""

from __future__ import annotations

import argparse
import json
import random
import re
import sys
import warnings
from collections import Counter
from pathlib import Path

warnings.filterwarnings("ignore")  # openpyxl is noisy about legacy workbook features

ROOT = Path(__file__).resolve().parents[1]
CORPUS = ROOT / "data" / "corpus"
RESULTS = ROOT / "results"

# Excel function call: NAME( -- uppercase, may contain digits/dots (e.g. LOG10, CEILING.MATH).
# Excludes anything preceded by an identifier char so cell refs and names do not match.
FUNC_CALL = re.compile(r"(?<![A-Za-z0-9_.$!])([A-Z][A-Z0-9.]{1,30})\s*\(")

# Not functions: Excel error literals and boolean constants that can appear bare.
NON_FUNCTIONS = {"TRUE", "FALSE"}


def supported_functions() -> set[str]:
    from xlcalculator import FUNCTIONS

    return {n.upper() for n in FUNCTIONS.keys() if not n.startswith("OP_")}


def functions_in(formula: str) -> set[str]:
    return {m.group(1) for m in FUNC_CALL.finditer(formula)} - NON_FUNCTIONS


def census(paths: list[Path], supported: set[str]) -> dict:
    from openpyxl import load_workbook

    per_function = Counter()          # function -> formula cells using it
    unsupported_hits = Counter()      # unsupported function -> formula cells using it
    wb_total = wb_readable = wb_no_formula = wb_fully_supported = 0
    cells_total = cells_supported = 0
    read_failures = Counter()

    for path in paths:
        wb_total += 1
        try:
            wb = load_workbook(path, data_only=False, read_only=True)
        except Exception as exc:  # noqa: BLE001 -- we are surveying, not executing
            read_failures[type(exc).__name__] += 1
            continue
        wb_readable += 1

        wb_cells = 0
        wb_ok = 0
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        v = cell.value
                        if not isinstance(v, str) or not v.startswith("="):
                            continue
                        wb_cells += 1
                        funcs = functions_in(v)
                        for f in funcs:
                            per_function[f] += 1
                        missing = funcs - supported
                        if missing:
                            for f in missing:
                                unsupported_hits[f] += 1
                        else:
                            wb_ok += 1
        except Exception as exc:  # noqa: BLE001
            read_failures[f"iter:{type(exc).__name__}"] += 1
        finally:
            wb.close()

        cells_total += wb_cells
        cells_supported += wb_ok
        if wb_cells == 0:
            wb_no_formula += 1
        elif wb_ok == wb_cells:
            wb_fully_supported += 1

    return {
        "workbooks_scanned": wb_total,
        "workbooks_readable": wb_readable,
        "workbooks_without_formulas": wb_no_formula,
        "workbooks_with_formulas": wb_readable - wb_no_formula,
        "workbooks_fully_supported": wb_fully_supported,
        "formula_cells": cells_total,
        "formula_cells_supported": cells_supported,
        "top_functions": per_function.most_common(30),
        "top_unsupported": unsupported_hits.most_common(30),
        "read_failures": dict(read_failures),
    }


def parse_check(paths: list[Path]) -> dict:
    """Does xlcalculator actually compile these workbooks?"""
    from xlcalculator import ModelCompiler

    ok = 0
    failures = Counter()
    for path in paths:
        try:
            ModelCompiler().read_and_parse_archive(str(path))
            ok += 1
        except Exception as exc:  # noqa: BLE001
            failures[type(exc).__name__] += 1
    return {"sampled": len(paths), "parsed": ok, "failures": dict(failures)}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=500, help="workbooks to survey")
    ap.add_argument("--parse-sample", type=int, default=0, help="workbooks to compile-test")
    ap.add_argument("--seed", type=int, default=17)
    args = ap.parse_args()

    if not CORPUS.exists():
        print("corpus missing -- run scripts/fetch_corpus.py first", file=sys.stderr)
        return 1

    all_paths = sorted(p for p in CORPUS.rglob("*.xlsx") if p.is_file())
    if not all_paths:
        print("no .xlsx files found under data/corpus", file=sys.stderr)
        return 1

    rng = random.Random(args.seed)
    paths = all_paths if len(all_paths) <= args.limit else rng.sample(all_paths, args.limit)

    supported = supported_functions()
    print(f"corpus holds {len(all_paths)} .xlsx files; surveying {len(paths)}")
    print(f"xlcalculator implements {len(supported)} functions\n")

    report = census(paths, supported)
    report["corpus_size"] = len(all_paths)
    report["seed"] = args.seed

    wf = report["workbooks_with_formulas"]
    print(f"readable workbooks          {report['workbooks_readable']}/{report['workbooks_scanned']}")
    print(f"  without any formulas      {report['workbooks_without_formulas']}")
    print(f"  with formulas             {wf}")
    if wf:
        pct = 100 * report["workbooks_fully_supported"] / wf
        print(f"  fully supported           {report['workbooks_fully_supported']} ({pct:.1f}%)")
    if report["formula_cells"]:
        pct = 100 * report["formula_cells_supported"] / report["formula_cells"]
        print(f"\nformula cells               {report['formula_cells']}")
        print(f"  evaluable                 {report['formula_cells_supported']} ({pct:.1f}%)")
        report["formula_cell_coverage_pct"] = round(pct, 2)

    print("\nmost used functions:")
    for name, n in report["top_functions"][:15]:
        mark = " " if name in supported else "X"
        print(f"  [{mark}] {name:16} {n:7d}")

    if report["top_unsupported"]:
        print("\nunsupported, by cost:")
        for name, n in report["top_unsupported"][:15]:
            print(f"      {name:16} {n:7d}")

    if report["read_failures"]:
        print(f"\nread failures: {report['read_failures']}")

    if args.parse_sample:
        sample = rng.sample(paths, min(args.parse_sample, len(paths)))
        print(f"\ncompile-testing {len(sample)} workbooks with xlcalculator …")
        pc = parse_check(sample)
        print(f"  parsed {pc['parsed']}/{pc['sampled']}")
        if pc["failures"]:
            print(f"  failures: {pc['failures']}")
        report["parse_check"] = pc

    RESULTS.mkdir(exist_ok=True)
    out = RESULTS / "coverage.json"
    out.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(f"\nwrote {out.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
