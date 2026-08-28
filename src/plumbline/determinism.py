"""Guard the one assumption every Plumbline proof depends on: recomputation is stable.

Plumbline proves a finding by evaluating a workbook twice -- once as-is, once with
a change applied -- and reporting the delta. That is only evidence if the workbook
returns the same numbers when nothing changes.

`RAND` appears in 45,550 cells of the Enron corpus (2.67%). xlcalculator supports
it, so it never surfaces as a coverage gap -- which makes it more dangerous, not
less. On such a workbook every evaluation differs, so a delta computed across two
runs is noise, and Plumbline would emit confident, precise, meaningless proofs.

Two defences, used together:

  `find_volatile`  -- static scan. Cheap, catches the known offenders by name.
  `check`          -- empirical. Evaluate twice, compare. Catches anything the
                      static scan missed, including volatility we have not thought of.

The static scan alone is not enough (an unknown volatile function slips through)
and the empirical check alone is not enough (RAND can coincidentally repeat on a
tiny sheet). Neither is sufficient; both together are.
"""

from __future__ import annotations

import re
import warnings
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

#: Excel functions that return a different value on every recalculation.
#: RAND/RANDBETWEEN are always unstable; NOW/TODAY are stable within a run but
#: drift across runs, which is the same problem on a slower clock.
VOLATILE_FUNCTIONS = frozenset({"RAND", "RANDBETWEEN", "NOW", "TODAY"})

_CALL = re.compile(r"(?<![A-Za-z0-9_.$!])([A-Z][A-Z0-9.]{1,30})\s*\(")


@dataclass
class VolatilityReport:
    """Where volatility lives in a workbook, and therefore what cannot be proved."""

    cells: dict[str, set[str]] = field(default_factory=dict)  # "Sheet!A1" -> {"RAND"}
    functions: set[str] = field(default_factory=set)

    @property
    def is_volatile(self) -> bool:
        return bool(self.cells)

    def __len__(self) -> int:
        return len(self.cells)

    def summary(self) -> str:
        if not self.cells:
            return "no volatile functions found"
        funcs = ", ".join(sorted(self.functions))
        return f"{len(self.cells)} volatile cell(s) using {funcs}"


def find_volatile(path: str | Path) -> VolatilityReport:
    """Static scan for volatile function calls. Fast, name-based, no evaluation."""
    report = VolatilityReport()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(path, data_only=False, read_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not isinstance(v, str) or not v.startswith("="):
                        continue
                    found = {m.group(1) for m in _CALL.finditer(v)} & VOLATILE_FUNCTIONS
                    if found:
                        report.cells[f"{ws.title}!{cell.coordinate}"] = found
                        report.functions |= found
    finally:
        wb.close()
    return report


@dataclass
class DeterminismResult:
    """Did two identical evaluations agree?"""

    stable: bool
    checked: int
    disagreements: dict[str, tuple] = field(default_factory=dict)
    error: str | None = None

    def summary(self) -> str:
        if self.error:
            return f"could not check: {self.error}"
        if self.stable:
            return f"stable across two runs ({self.checked} cells compared)"
        n = len(self.disagreements)
        sample = list(self.disagreements.items())[:3]
        detail = "; ".join(f"{ref} {a} vs {b}" for ref, (a, b) in sample)
        return f"UNSTABLE: {n}/{self.checked} cells disagreed ({detail})"


def check(path: str | Path, limit: int = 400) -> DeterminismResult:
    """Evaluate the workbook twice and compare. The empirical half of the guard.

    Two full parses, deliberately: reusing one parsed model would share cached
    values and could report stability that does not exist.
    """
    from xlcalculator import Evaluator, ModelCompiler

    def evaluate_all() -> dict[str, object]:
        model = ModelCompiler().read_and_parse_archive(str(path))
        ev = Evaluator(model)
        out: dict[str, object] = {}
        for addr, cell in model.cells.items():
            if not getattr(cell.formula, "formula", None):
                continue
            if len(out) >= limit:
                break
            try:
                out[addr] = _native(ev.evaluate(addr))
            except Exception as exc:  # noqa: BLE001 -- a cell that errors is still deterministic
                out[addr] = f"<error:{type(exc).__name__}>"
        return out

    try:
        first = evaluate_all()
        second = evaluate_all()
    except Exception as exc:  # noqa: BLE001
        return DeterminismResult(stable=False, checked=0, error=f"{type(exc).__name__}: {exc}")

    disagreements = {
        ref: (first[ref], second[ref]) for ref in first if first[ref] != second.get(ref)
    }
    return DeterminismResult(
        stable=not disagreements, checked=len(first), disagreements=disagreements
    )


def freeze(path: str | Path, dest: str | Path) -> VolatilityReport:
    """Write a copy with volatile formulas replaced by their last cached values.

    Excel stores the value from its most recent calculation alongside the formula.
    Substituting that value makes the workbook stable, so the rest of the model can
    be audited. The frozen cells themselves become constants and are excluded from
    audit -- we can no longer say anything about them, and pretending otherwise
    would be the same silent wrongness we are guarding against.
    """
    report = find_volatile(path)
    if not report.is_volatile:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            load_workbook(path).save(dest)
        return report

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        cached = load_workbook(path, data_only=True)   # values from Excel's last calc
        live = load_workbook(path, data_only=False)    # formulas

    for ref in report.cells:
        sheet, _, coord = ref.partition("!")
        value = cached[sheet][coord].value
        # A workbook never opened in Excel has no cached values; 0 would be a lie,
        # so leave the formula in place and let the determinism check reject it.
        if value is not None:
            live[sheet][coord] = value

    live.save(dest)
    return report


def _native(value):
    """xlcalculator returns its own numeric wrappers; unwrap for comparison."""
    if isinstance(value, bool) or value is None:
        return value
    try:
        f = float(value)
    except (TypeError, ValueError):
        return str(value)
    return round(f, 10)  # kill float jitter that is not real instability
