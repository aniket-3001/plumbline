"""Tests for Layer 2 -- the only place a model is allowed to speak.

The rules being tested here are the fence around that model. Without them a
hallucinated cell reference reaches an analyst's audit report, which is worse than
saying nothing: it sends them chasing a cell that was never wrong.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

import plumbline  # noqa: F401
from plumbline.agent import (
    SYSTEM_PROMPT,
    Interpretation,
    _render_user_prompt,
    _strip_fence,
    build_context,
    interpret,
)
from plumbline.audit import audit

ROOT = Path(__file__).resolve().parents[1]
CLEAN = ROOT / "tests" / "fixtures" / "quarterly_pl.xlsx"


@pytest.fixture(scope="module")
def finding():
    return audit(CLEAN, check_determinism=False).proved[0]


def stub(payload) -> callable:
    text = payload if isinstance(payload, str) else json.dumps(payload)
    return lambda system, user: text


class TestContext:
    def test_finds_the_labels_that_give_the_cell_meaning(self, finding):
        ctx = build_context(str(CLEAN), finding)
        assert ctx["row_label"] == "Total Opex"
        assert ctx["column_label"] == "Q2"

    def test_includes_row_peers(self, finding):
        ctx = build_context(str(CLEAN), finding)
        peers = {p["cell"] for p in ctx["row_peers"]}
        assert {"B11", "D11", "E11"} <= peers

    def test_excludes_the_cell_itself_from_peers(self, finding):
        ctx = build_context(str(CLEAN), finding)
        assert finding.cell not in {p["cell"] for p in ctx["row_peers"]}

    def test_stays_small(self, finding):
        """A whole sheet neither fits a context window nor helps."""
        prompt = _render_user_prompt(build_context(str(CLEAN), finding))
        assert len(prompt) < 2000

    def test_carries_the_proof_so_the_model_does_not_relitigate_it(self, finding):
        ctx = build_context(str(CLEAN), finding)
        assert "27000" in ctx["proof"] and "30000" in ctx["proof"]


class TestHallucinationGuard:
    def test_accepts_cells_present_in_the_context(self, finding):
        result = interpret(str(CLEAN), finding, stub({
            "intent": "Total operating expenses for Q2",
            "deliberate": False,
            "explanation": "Every other quarter sums three lines; Q2 sums two.",
            "cells_referenced": ["C8", "C10", "B11"],
        }))
        assert result.ok
        assert result.rejected_cells == []

    def test_rejects_an_invented_cell(self, finding):
        result = interpret(str(CLEAN), finding, stub({
            "intent": "x", "deliberate": False, "explanation": "y",
            "cells_referenced": ["ZZ999"],
        }))
        assert not result.ok
        assert result.rejected_cells == ["ZZ999"]
        assert "ZZ999" in result.error

    def test_one_bad_reference_invalidates_the_whole_interpretation(self, finding):
        """Partial trust is not a thing: if it invented one cell, drop the claim."""
        result = interpret(str(CLEAN), finding, stub({
            "intent": "x", "deliberate": False, "explanation": "y",
            "cells_referenced": ["C8", "QQ42"],
        }))
        assert not result.ok


class TestRobustness:
    def test_survives_a_model_outage(self, finding):
        def boom(system, user):
            raise ConnectionError("upstream down")

        result = interpret(str(CLEAN), finding, boom)
        assert not result.ok
        assert "ConnectionError" in result.error

    def test_survives_unparsable_output(self, finding):
        result = interpret(str(CLEAN), finding, stub("not json at all"))
        assert not result.ok
        assert "unparsable" in result.error

    def test_strips_markdown_fences(self):
        assert _strip_fence('```json\n{"a": 1}\n```') == '{"a": 1}'
        assert _strip_fence('{"a": 1}') == '{"a": 1}'

    def test_truncates_runaway_output(self, finding):
        result = interpret(str(CLEAN), finding, stub({
            "intent": "x" * 5000, "deliberate": None,
            "explanation": "y" * 5000, "cells_referenced": [],
        }))
        assert len(result.intent) <= 400
        assert len(result.explanation) <= 800


class TestPromptContract:
    def test_forbids_overturning_the_proof(self):
        """Recomputation decides errors. The model supplies intent only."""
        assert "NOT deciding whether this is an error" in SYSTEM_PROMPT

    def test_forbids_inventing_references(self):
        assert "Never invent a cell reference" in SYSTEM_PROMPT.replace("\\n", "")

    def test_permits_admitting_ignorance(self):
        """A model with no way to say 'unclear' will confabulate instead."""
        assert "do not indicate the intent" in SYSTEM_PROMPT


def test_interpretation_serialises():
    assert Interpretation(intent="i").to_dict()["intent"] == "i"


class TestLabelExtraction:
    """A label is what a reader sees, never the formula text behind it."""

    @staticmethod
    def _finding(sheet="S", cell="C5"):
        return type(
            "F", (), {
                "sheet": sheet, "cell": cell,
                "actual": "=B4", "expected": "=B5",
                "proof": "C5: 1 -> 2 (+1)",
            },
        )()

    def test_a_computed_header_is_not_passed_through_as_a_formula(self, tmp_path):
        """Financial models label columns with computed dates. Reading the first
        string in the column returns `=B1+1`, which is not a label, is not what
        anyone sees on screen, and invites the model to reason about a header that
        does not exist."""
        from openpyxl import Workbook

        from plumbline.agent import build_context

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "start"
        ws["B1"] = 1
        ws["C1"] = "=B1+1"          # computed header, no cached value
        ws["A5"] = "Revenue"
        ws["B5"] = 10
        ws["C5"] = "=B4"
        path = tmp_path / "computed_header.xlsx"
        wb.save(path)

        ctx = build_context(str(path), self._finding())
        assert ctx["column_label"] != "=B1+1"
        assert ctx["column_label"] is None or not str(ctx["column_label"]).startswith("=")

    def test_a_real_text_header_is_still_found(self, tmp_path):
        from openpyxl import Workbook

        from plumbline.agent import build_context

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["C1"] = "Q3 2002"
        ws["A5"] = "Revenue"
        ws["B5"] = 10
        ws["C5"] = "=B4"
        path = tmp_path / "text_header.xlsx"
        wb.save(path)

        ctx = build_context(str(path), self._finding())
        assert ctx["column_label"] == "Q3 2002"
        assert ctx["row_label"] == "Revenue"

    def test_the_prompt_says_none_found_rather_than_inventing_one(self, tmp_path):
        from openpyxl import Workbook

        from plumbline.agent import _render_user_prompt, build_context

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["B5"] = 10
        ws["C5"] = "=B4"
        path = tmp_path / "no_labels.xlsx"
        wb.save(path)

        prompt = _render_user_prompt(build_context(str(path), self._finding()))
        assert "(none found)" in prompt


class TestGuardScope:
    """The guard's idea of "the context" must match the model's: the rendered prompt.

    Found by replaying a real trajectory, not by reasoning about it. On
    `chris_germany__1938!U8` the prompt lists the peer `Q8: =+P8`; a correct answer
    cited P8 and the guard rejected it as a hallucination. A guard that punishes
    correct reasoning gets switched off, and then it protects nothing.
    """

    @staticmethod
    def _ctx():
        return {
            "sheet": "Sheet1", "cell": "U8",
            "row_label": "Fixed", "column_label": None,
            "actual_formula": "=+T7", "expected_formula": "=+T8",
            "row_peers": [
                {"cell": "Q8", "formula": "=+P8", "value": 1},
                {"cell": "T8", "formula": "=+S8", "value": 2},
            ],
            "precedents": [{"cell": "T8", "formula": "=+S8", "value": 2}],
            "proof": "U8: 10000 -> 2.1562 (-9997.8438)",
        }

    def test_a_cell_named_inside_a_shown_peer_formula_is_known(self):
        from plumbline.agent import _known_cells

        known = _known_cells(self._ctx())
        assert "P8" in known, "P8 is printed in the prompt as `Q8: =+P8`"
        assert {"U8", "T7", "T8", "S8", "Q8"} <= known

    def test_a_cell_never_shown_is_still_rejected(self):
        from plumbline.agent import _known_cells

        assert "AJ40" not in _known_cells(self._ctx())

    def test_a_cross_sheet_reference_is_still_rejected(self):
        import json

        from plumbline.agent import interpret

        reply = json.dumps({
            "intent": "x", "deliberate": True, "explanation": "y",
            "cells_referenced": ["U8", "Summary!B12"],
        })
        finding = type("F", (), {
            "sheet": "Sheet1", "cell": "U8", "actual": "=+T7",
            "expected": "=+T8", "proof": "U8: 1 -> 2 (+1)",
        })()

        import plumbline.agent as agent

        original = agent.build_context
        agent.build_context = lambda *a, **k: self._ctx()
        try:
            interp = interpret("ignored.xlsx", finding, lambda s, u: reply)
        finally:
            agent.build_context = original

        assert interp.ok is False
        assert "SUMMARY!B12" in interp.rejected_cells
