"""Tests for the error-seeding harness.

These matter more than they look. The seeded workbooks *are* the ground truth, so
a bug here does not fail loudly -- it silently produces a benchmark that measures
something other than what we claim.
"""

from __future__ import annotations

import random
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import plumbline  # noqa: F401,E402
from plumbline.seeding import (  # noqa: E402
    Seed,
    _majority_rows,
    _shift_first_reference,
    _shift_range_end,
    _swap_function,
    seed_workbook,
)

CLEAN = ROOT / "tests" / "fixtures" / "quarterly_pl.xlsx"


class TestMutations:
    def test_shift_range_end_drops_one_row(self):
        assert _shift_range_end("=SUM(B8:B10)", -1) == "=SUM(B8:B9)"

    def test_shift_range_end_returns_none_without_a_range(self):
        assert _shift_range_end("=B8*2", -1) is None

    def test_shift_range_end_refuses_to_go_above_row_one(self):
        assert _shift_range_end("=SUM(A1:A1)", -1) is None

    def test_shift_reference_moves_a_plain_reference(self):
        assert _shift_first_reference("=D19", -1) == "=D18"
        assert _shift_first_reference("=D19*2", -1) == "=D18*2"

    def test_shift_reference_leaves_ranges_alone(self):
        """Ranges belong to the other mutation, so the classes stay distinct."""
        assert _shift_first_reference("=SUM(B8:B10)", -1) is None

    def test_shift_reference_respects_absolute_anchors(self):
        """A $-anchored row was anchored on purpose; moving it is a different error."""
        assert _shift_first_reference("=D$19", -1) is None

    def test_swap_function(self):
        assert _swap_function("=SUM(B8:B10)", "SUM", "AVERAGE") == "=AVERAGE(B8:B10)"
        assert _swap_function("=B8-B9", "SUM", "AVERAGE") is None


class TestMajorityRows:
    def test_requires_unanimity(self):
        """A row containing a deviation is not a safe place to seed.

        Row 11 of the clean fixture holds the original off-by-one, so it must not
        be offered as a seeding site -- the majority has to survive the seed.
        """
        from poc import load_formulas

        rows = _majority_rows(load_formulas(str(CLEAN))["P&L"])
        assert 5 in rows and 13 in rows   # unanimous rows
        assert 11 not in rows             # already contains C11's deviation

    def test_ignores_rows_with_too_few_members(self):
        assert _majority_rows({"A1": "=B1", "A2": "=B2"}) == {}


class TestDifficulty:
    @pytest.mark.parametrize(
        "panko,changed,value,expected",
        [
            ("hardcoding", False, 100, "silent"),
            ("mechanical", True, 0, "obvious"),
            ("mechanical", True, None, "obvious"),
            ("mechanical", True, "<error:X>", "obvious"),
            ("mechanical", True, 45.75, "realistic"),
            ("logic", True, 1234, "realistic"),
        ],
    )
    def test_classification(self, panko, changed, value, expected):
        seed = Seed(
            seed_id="t",
            panko_class=panko,
            sheet="S",
            cell="A1",
            original_formula="=B1",
            seeded_formula="=B2",
            description="",
            detectable_by=[],
            seeded_value=value,
            value_changed=changed,
        )
        assert seed.classify_difficulty() == expected


class TestSeedWorkbook:
    def test_produces_a_manifest_and_a_file(self, tmp_path):
        manifest = seed_workbook(CLEAN, tmp_path, random.Random(7), max_seeds=3)
        assert manifest is not None
        assert (tmp_path / CLEAN.name).exists()
        assert manifest["seed_count"] == len(manifest["seeds"]) > 0

    def test_records_pre_existing_findings_separately(self, tmp_path):
        """The clean fixture already contains C11. A detector hitting it is not a
        false positive, and conflating the two would understate precision."""
        manifest = seed_workbook(CLEAN, tmp_path, random.Random(7), max_seeds=3)
        assert "P&L!C11" in manifest["pre_existing_findings"]

    def test_every_seed_is_verified_to_have_done_something(self, tmp_path):
        """A seed that changes nothing is noise in the ground truth, not an error.

        Hardcoding is the deliberate exception: it is defined by being identical
        today and only diverging later.
        """
        manifest = seed_workbook(CLEAN, tmp_path, random.Random(7), max_seeds=3)
        for seed in manifest["seeds"]:
            if seed["panko_class"] == "hardcoding":
                assert seed["value_changed"] is False
                assert seed["difficulty"] == "silent"
            else:
                assert seed["value_changed"] is True

    def test_seeded_file_matches_the_manifest(self, tmp_path):
        """The file on disk must contain exactly what the manifest claims."""
        from openpyxl import load_workbook

        manifest = seed_workbook(CLEAN, tmp_path, random.Random(7), max_seeds=3)
        wb = load_workbook(tmp_path / CLEAN.name)
        for seed in manifest["seeds"]:
            actual = wb[seed["sheet"]][seed["cell"]].value
            expected = seed["seeded_formula"]
            if isinstance(expected, str):
                assert actual == expected
            else:
                assert float(actual) == pytest.approx(float(expected))

    def test_is_deterministic_for_a_given_seed(self, tmp_path):
        """Same RNG seed, same injected errors -- or the evaluation is not reproducible."""
        a = seed_workbook(CLEAN, tmp_path / "a", random.Random(11), max_seeds=3)
        b = seed_workbook(CLEAN, tmp_path / "b", random.Random(11), max_seeds=3)
        assert [(s["cell"], s["panko_class"]) for s in a["seeds"]] == [
            (s["cell"], s["panko_class"]) for s in b["seeds"]
        ]


class TestPreExistingFindings:
    """The benchmark's own accounting, which is easy to get wrong quietly."""

    @staticmethod
    def _messy_workbook(path):
        """A row of `=Bn+1` counters with a typed constant spliced in -- the exact
        shape of the real hardcoding found across the Enron corpus."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = 1
        for col in range(2, 9):
            ws.cell(row=1, column=col, value=f"={chr(64 + col - 1)}1+1")
        ws["E1"] = 5          # dead cell: the right value, but typed, not computed
        ws["A3"] = 10
        for col in range(2, 9):
            ws.cell(row=3, column=col, value=f"=SUM($A$3:{chr(64 + col - 1)}3)")
        ws["F3"] = "=SUM($A$3:D3)"   # pattern break
        wb.save(path)
        return path

    def test_finds_dead_cells_not_only_pattern_breaks(self, tmp_path):
        from plumbline.seeding import pre_existing_findings

        path = self._messy_workbook(tmp_path / "messy.xlsx")
        found = pre_existing_findings(path)
        # Both detectors must contribute. Running only the pattern-break detector
        # made every pre-existing dead cell surface later as a false positive.
        assert "S!E1" in found, f"dead cell missed: {found}"
        assert "S!F3" in found, f"pattern break missed: {found}"

    def test_agrees_cell_for_cell_with_what_the_audit_reports(self, tmp_path):
        from plumbline.audit import audit
        from plumbline.seeding import pre_existing_findings

        path = self._messy_workbook(tmp_path / "messy2.xlsx")
        report = audit(path, check_determinism=False)
        assert set(pre_existing_findings(path)) == {
            f"{f.sheet}!{f.cell}" for f in report.findings
        }

    def test_a_clean_workbook_has_nothing_pre_existing(self, tmp_path):
        from openpyxl import Workbook

        from plumbline.seeding import pre_existing_findings

        wb = Workbook()
        ws = wb.active
        ws["A1"] = 5
        for col in range(2, 8):
            ws.cell(row=1, column=col, value=f"={chr(64 + col - 1)}1+1")
        path = tmp_path / "clean.xlsx"
        wb.save(path)
        assert pre_existing_findings(path) == []


class TestSeedsNeverFlipTheMajority:
    def test_at_most_one_seed_per_row(self, tmp_path):
        """Two seeds in one row can make the correct cells the minority, so the
        benchmark would be scoring the tool for finding the untouched cell."""
        import random

        from openpyxl import Workbook

        from plumbline.seeding import plan_seeds

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        for row in range(1, 6):
            ws.cell(row=row, column=1, value=row)
            for col in range(2, 6):
                ws.cell(row=row, column=col, value=f"=SUM($A${row}:{chr(64 + col - 1)}{row})")
        path = tmp_path / "rows.xlsx"
        wb.save(path)

        plan = plan_seeds(path, random.Random(0), max_seeds=8)
        rows = [(c["sheet"], "".join(ch for ch in c["cell"] if ch.isdigit())) for c in plan]
        assert len(rows) == len(set(rows)), f"two seeds landed in one row: {plan}"


class TestExclusionsTrackDetectorSettings:
    """The exclusion list must be computed at the same sensitivity as the audit.

    Raise the audit's sensitivity without raising the exclusion list's, and every
    extra pre-existing cell the audit now correctly finds is charged to it as a
    false positive -- so a detector improvement reads as a precision collapse.
    """

    @staticmethod
    def _two_peer_row(path):
        """A row with exactly two formula peers and a typed constant between them.
        Visible at min_peers=2, invisible at min_peers=3."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        # Detection is row-wise, so the peers must sit in the same row as the cell.
        for col, value in (("B", 5), ("C", 10), ("D", 15)):
            ws[f"{col}1"] = value
        ws["B2"] = "=B1*2"
        ws["C2"] = 20          # dead cell: equals C1*2, but typed, not computed
        ws["D2"] = "=D1*2"
        wb.save(path)
        return path

    def test_a_lower_threshold_sees_more(self, tmp_path):
        from plumbline.seeding import pre_existing_findings

        path = self._two_peer_row(tmp_path / "two_peers.xlsx")
        assert "S!C2" in pre_existing_findings(path, min_peers=2)
        assert "S!C2" not in pre_existing_findings(path, min_peers=3)

    def test_matched_settings_agree_with_the_audit(self, tmp_path):
        from plumbline.audit import audit
        from plumbline.seeding import pre_existing_findings

        path = self._two_peer_row(tmp_path / "matched.xlsx")
        for peers in (2, 3):
            report = audit(path, check_determinism=False, min_peers=peers)
            assert set(pre_existing_findings(path, min_peers=peers)) == {
                f"{f.sheet}!{f.cell}" for f in report.findings
            }, f"disagreement at min_peers={peers}"

    def test_mismatched_settings_are_what_produce_phantom_false_positives(self, tmp_path):
        """Documents the failure mode directly: the audit finds a cell the
        exclusion list, computed at a stricter threshold, does not carry."""
        from plumbline.audit import audit
        from plumbline.seeding import pre_existing_findings

        path = self._two_peer_row(tmp_path / "mismatched.xlsx")
        found = {f"{f.sheet}!{f.cell}" for f in audit(path, check_determinism=False, min_peers=2).findings}
        stale = set(pre_existing_findings(path, min_peers=3))
        assert found - stale == {"S!C2"}


class TestContiguity:
    """Peers must sit next to the candidate, not merely somewhere in its row.

    The `min_peers` ablation hand-labelled 29 findings and found every clear false
    positive was a cross-block comparison: candidate in one block of the row, peers
    in another, data or blanks between. These fixtures are the shapes of the two
    real ones, reduced to their essentials.
    """

    @staticmethod
    def _build(path, cells):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        for ref, value in cells.items():
            ws[ref] = value
        wb.save(path)
        return path

    @staticmethod
    def _dead(path, *, contiguous, min_peers=2):
        from plumbline.audit import detect_dead_cells, screen_dead_cells
        from poc import load_formulas

        out = []
        for sheet, formulas in load_formulas(str(path)).items():
            out += detect_dead_cells(str(path), sheet, formulas,
                                     min_peers=min_peers, contiguous=contiguous)
        return {f"{f.sheet}!{f.cell}" for f in screen_dead_cells(str(path), out)}

    def test_a_lookup_table_data_column_is_not_a_dead_formula(self, tmp_path):
        """`MANUAL` row 28 of ONEOKRECAP2001: two `(counter, name, id)` blocks side by
        side. `B` and `G` are the counters; `D` is a meter-ID column that happens to
        run consecutively, so it passes the value screen by coincidence."""
        path = self._build(tmp_path / "lookup.xlsx", {
            "B26": 1, "D26": 50, "G26": 1,
            "B27": "=B26+1", "C27": "FIN 1", "D27": 51, "G27": "=G26+1", "H27": "LIBERAL",
            "B28": "=B27+1", "C28": "FIN 2", "D28": 52, "G28": "=G27+1", "H28": "STC 1",
        })
        assert "S!D28" in self._dead(path, contiguous=False), "fixture must reproduce the bug"
        assert "S!D28" not in self._dead(path, contiguous=True)

    def test_a_price_in_a_data_block_is_not_a_dead_formula(self, tmp_path):
        """`Options` row 16: `(B,C,D)` carry forward; `(E,F,G)` are strike-price data.
        `F16` is data, and `B16`/`D16` are in a different block."""
        path = self._build(tmp_path / "options.xlsx", {
            "B15": 10, "D15": 20,
            "B16": "=B15", "C16": "Dec 01", "D16": "=D15",
            "E16": "40 mp", "F16": 1, "G16": 1.5,
        })
        assert "S!F16" not in self._dead(path, contiguous=True)

    def test_a_real_frozen_formula_inside_a_run_still_reports(self, tmp_path):
        """The shape of `chris_germany__1938!AH25`: an unbroken carry-forward run with
        one cell typed over. Contiguity must not cost this."""
        path = self._build(tmp_path / "frozen.xlsx", {
            "B25": 5000,
            "C25": "=B25", "D25": "=C25", "E25": "=D25",
            "F25": 5000,          # frozen: right value, no formula
            "G25": "=F25",
        })
        assert "S!F25" in self._dead(path, contiguous=True)

    def test_a_gap_is_a_block_boundary(self, tmp_path):
        """An empty column between candidate and peers is a boundary, not something
        to step over -- that gap is what separates two blocks in a real sheet."""
        path = self._build(tmp_path / "gap.xlsx", {
            "A1": 1, "B1": "=A1+1", "C1": "=B1+1",
            "D1": 3, "E1": 4,          # `=D1+1` really is 4: only contiguity rejects it
            "G1": 6, "H1": "=G1+1", "I1": "=H1+1",
        })
        assert "S!E1" not in self._dead(path, contiguous=True)


class TestContiguityBlindSpot:
    """Contiguity assumes the corruption is one cell wide. Sometimes it is not.

    Both shapes below are real, from `scott_neal__38672`, and both are cells this
    detector genuinely should flag and does not. They are pinned as tests so the
    limitation is a known quantity rather than a surprise, and so that any future
    fix has a target to turn green.
    """

    @staticmethod
    def _dead(path):
        from plumbline.audit import detect_dead_cells, screen_dead_cells
        from poc import load_formulas

        out = []
        for sheet, formulas in load_formulas(str(path)).items():
            out += detect_dead_cells(str(path), sheet, formulas)
        return {f"{f.sheet}!{f.cell}" for f in screen_dead_cells(str(path), out)}

    @staticmethod
    def _save(path, cells):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        for ref, value in cells.items():
            ws[ref] = value
        wb.save(path)
        return path

    def test_two_adjacent_overwrites_hide_each_other(self, tmp_path):
        """`Floor Plan` row 32: `R S T` are `=<next>+1`, then U=603, V=602, W=601.
        Sibling blocks in the same row carry exactly one trailing constant, so U and
        V were both formulas once. The outer one is still found; the inner one is
        not, because its neighbour is no longer a formula either."""
        path = self._save(tmp_path / "wide.xlsx", {
            "R1": "=S1+1", "S1": "=T1+1", "T1": "=U1+1",
            "U1": 603, "V1": 602, "W1": 601,
        })
        found = self._dead(path)
        assert "S!U1" in found, "the overwrite adjacent to live formulas is still caught"
        assert "S!V1" not in found, "known blind spot: its only neighbour is also dead"

    def test_a_block_with_no_surviving_formula_is_unreachable(self, tmp_path):
        """`Floor Plan!Y41`/`Y64`: the whole two-cell block was overwritten, so
        nothing in the row can vouch for it. No layout rule can recover this one --
        it would need the labels."""
        path = self._save(tmp_path / "allgone.xlsx", {
            "D1": "=E1+1", "E1": "=F1+1", "F1": 527,
            "H1": 212, "I1": 211,
        })
        assert "S!H1" not in self._dead(path)

    def test_the_benchmark_cannot_see_this(self):
        """Stated as an executable note: the seeder overwrites exactly one cell per
        row, so every seeded dead cell is the case contiguity handles best. The
        recall gain it measures is therefore an over-estimate for workbooks whose
        corruption is wider, and the two tests above are the evidence."""
        from plumbline.seeding import plan_seeds  # noqa: F401  -- one seed per row

        assert True
