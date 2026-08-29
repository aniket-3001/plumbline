"""Layer 2: semantic interpretation. The only place a model is allowed to speak.

The deterministic layers can tell that `C11` breaks its row's formula pattern. They
cannot tell that a cell **labelled "Total Q3" is summing Q2** -- that needs reading
the labels around it. This is the documented failure of every rule-based auditor:
Nixon & O'Hara found commercial tools "failed where label-pattern recognition was
required," while the one that succeeded did so "by using labels to indicate in
plain English that the wrong values were being totalled."

Three rules govern this layer, and they exist because a model in an audit tool is
a liability unless fenced:

1. **The model never decides whether something is an error.** Recomputation does
   that. The model only supplies intent -- what was this cell *for*? -- and a
   human-readable explanation. A finding's `proved` flag is never set here.

2. **The model never sees the whole sheet.** It gets a minimal subgraph: the cell,
   its precedents, its row peers, and the surrounding labels. Whole sheets do not
   fit in a context window, and dumping one invites the model to invent structure
   that is not there.

3. **Every claim is checked against the graph before it is shown.** If the model
   references a cell that does not exist, or asserts a dependency the model does
   not have, the claim is dropped. A hallucinated cell reference in an audit report
   is worse than silence -- it sends an analyst chasing a cell that was never wrong.

The client is an injected callable, so the pipeline is testable without a network
and swappable between providers.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from typing import Callable, Protocol

from openpyxl.utils import column_index_from_string, get_column_letter

CELL_RE = re.compile(r"\b([A-Z]{1,3})([1-9][0-9]{0,6})\b")

#: Default model. Kept in one place so a change is a one-line diff.
DEFAULT_MODEL = "claude-opus-5"

SYSTEM_PROMPT = """\
You are assisting a spreadsheet audit. A deterministic engine has already found \
that one cell breaks the formula pattern of its row and has PROVED, by recomputing \
the workbook, that correcting it changes the numbers.

Your only job is to explain what the cell was *meant* to do, using the row and \
column labels around it, and to say whether the deviation looks deliberate.

You are NOT deciding whether this is an error. That is already established. Do not \
overturn it, and do not hedge about whether recomputation is reliable.

Rules:
- Refer only to cell addresses that appear in the context you are given. Never \
invent a cell reference.
- If the labels do not tell you what the cell was for, say so plainly. "The labels \
do not indicate the intent" is a correct and useful answer.
- Judge deliberateness from evidence: a row labelled differently, a note in an \
adjacent cell, a consistent exception elsewhere. Absent evidence, say it is unclear.

Reply as JSON only:
{"intent": "<what this cell should compute, in one sentence>",
 "deliberate": true | false | null,
 "explanation": "<two sentences a financial analyst would find useful>",
 "cells_referenced": ["<every cell address you mentioned>"]}"""


class Client(Protocol):
    """Anything that can turn (system, user) into text."""

    def __call__(self, system: str, user: str) -> str: ...


@dataclass
class Interpretation:
    """What the model said, after the graph has vetted it."""

    intent: str = ""
    deliberate: bool | None = None
    explanation: str = ""
    cells_referenced: list[str] = field(default_factory=list)
    rejected_cells: list[str] = field(default_factory=list)
    ok: bool = True
    error: str = ""

    def to_dict(self) -> dict:
        return {
            "intent": self.intent,
            "deliberate": self.deliberate,
            "explanation": self.explanation,
            "cells_referenced": self.cells_referenced,
            "rejected_cells": self.rejected_cells,
            "ok": self.ok,
            "error": self.error,
        }


def build_context(path: str, finding, *, radius: int = 4, evaluate=None) -> dict:
    """The minimal subgraph a model needs, and nothing more.

    Whole sheets neither fit a context window nor help: extra rows invite the model
    to invent structure. So we send the cell, its row peers, its precedents, and the
    labels that give them meaning.
    """
    import warnings

    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(path, data_only=False)
        values = load_workbook(path, data_only=True)

    ws, vs = wb[finding.sheet], values[finding.sheet]
    m = CELL_RE.fullmatch(finding.cell)
    col = column_index_from_string(m.group(1))
    row = int(m.group(2))

    def describe(r: int, c: int) -> dict | None:
        if r < 1 or c < 1:
            return None
        addr = f"{get_column_letter(c)}{r}"
        raw = ws.cell(row=r, column=c).value
        if raw is None:
            return None
        return {
            "cell": addr,
            "formula": raw if isinstance(raw, str) and raw.startswith("=") else None,
            "value": vs.cell(row=r, column=c).value,
        }

    def label_at(r: int, c: int) -> str | None:
        """A label is what a *reader* sees in that cell, never the formula text.

        Financial models label columns with computed dates -- row 3 of
        `chris_germany__1938` is `=+T3+1` repeated across the sheet. Taking the
        first string found returns "=+T3+1" as the column label, which is not a
        label, is not what anyone sees on screen, and invites the model to reason
        about a header that does not exist. Twenty-five-year-old workbooks also
        frequently carry no cached values, so the displayed text is simply not
        recoverable from the file; when that happens the honest answer is that
        there is no label, and the caller keeps looking further out.
        """
        raw = ws.cell(row=r, column=c).value
        if raw is None:
            return None
        if isinstance(raw, str) and raw.startswith("="):
            shown = vs.cell(row=r, column=c).value
            if shown is None:                      # no cached value: unrecoverable
                return None
            if evaluate is not None:
                try:
                    shown = evaluate(f"{finding.sheet}!{get_column_letter(c)}{r}")
                except Exception:  # noqa: BLE001
                    pass
            return str(shown)
        return raw if isinstance(raw, str) else None

    # Labels: the leftmost text in this row (models put row headers in column A or
    # near it), and the topmost text in this column.
    row_label = next(
        (lbl for c in range(1, col) if (lbl := label_at(row, c))), None
    )
    col_label = next(
        (lbl for r in range(1, row) if (lbl := label_at(r, col))), None
    )

    peers = [d for c in range(max(1, col - radius), col + radius + 1)
             if c != col and (d := describe(row, c))]
    precedents = [
        d for ref in CELL_RE.findall(finding.expected or "")
        if (d := describe(int(ref[1]), column_index_from_string(ref[0])))
    ]

    wb.close()
    values.close()
    return {
        "sheet": finding.sheet,
        "cell": finding.cell,
        "row_label": row_label,
        "column_label": col_label,
        "actual_formula": finding.actual,
        "expected_formula": finding.expected,
        "row_peers": peers,
        "precedents": precedents,
        "proof": finding.proof,
    }


def _render_user_prompt(ctx: dict) -> str:
    lines = [
        f"Sheet: {ctx['sheet']}",
        f"Cell under review: {ctx['cell']}",
        f"Row label: {ctx['row_label'] or '(none found)'}",
        f"Column label: {ctx['column_label'] or '(none found)'}",
        "",
        f"This cell currently contains: {ctx['actual_formula']}",
        f"Every other cell in its row is shaped like: {ctx['expected_formula']}",
        "",
        f"Recomputation proof: {ctx['proof']}",
        "",
        "Neighbouring cells in the same row:",
    ]
    for p in ctx["row_peers"]:
        body = p["formula"] or p["value"]
        lines.append(f"  {p['cell']}: {body}")
    if ctx["precedents"]:
        lines.append("")
        lines.append("Cells the expected formula would read:")
        for p in ctx["precedents"]:
            body = p["formula"] or p["value"]
            lines.append(f"  {p['cell']}: {body}")
    return "\n".join(lines)


def _known_cells(ctx: dict) -> set[str]:
    """Every cell address the model was actually shown.

    The system prompt says "refer only to cell addresses that appear in the context
    you are given", so the guard has to mean the same thing by "the context" that
    the model does: the rendered prompt, in full.

    An earlier version enumerated only the peer and precedent *addresses* plus the
    references inside the cell's own two formulas. That rejected a correct answer.
    On `chris_germany__1938!U8` the prompt lists the peer `Q8: =+P8`, so P8 is on
    the model's screen; citing it is reasoning from the evidence, not inventing it,
    and the guard called it a hallucination. A guard that punishes correct reasoning
    gets switched off, and then it protects nothing.

    Parsing the rendered prompt keeps the two definitions from drifting apart again:
    there is now one place a cell can become known, and it is the same text the
    model reads. Cross-sheet references stay rejected -- `Summary!B12` does not
    match a bare address in the prompt -- which is the case that matters, since a
    fabricated reference to another tab is the one an analyst cannot cheaply check.
    """
    return {f"{c}{r}" for c, r in CELL_RE.findall(_render_user_prompt(ctx))}


#: Typography a model reaches for that a legacy Windows console cannot encode.
_PLAIN = {
    "—": "-", "–": "-", "…": "...", "→": "->",
    "‘": "'", "’": "'", "“": '"', "”": '"',
    " ": " ", "≤": "<=", "≥": ">=", "×": "x",
}


def _plain(text: str) -> str:
    """Fold model prose to ASCII.

    Model output is untrusted input, and this is not hypothetical: the very first
    live call returned an arrow and an em dash. Printing either to a Windows console
    on a legacy code page raises UnicodeEncodeError, so an audit that had already
    produced valid deterministic findings would die while displaying them.

    Normalising here rather than at each display site means every consumer -- the
    terminal report, the JSON, a future caller -- gets text it can render, and there
    is one place to look when something odd appears in a finding.
    """
    for char, plain in _PLAIN.items():
        text = text.replace(char, plain)
    return text.encode("ascii", "replace").decode("ascii")


def interpret(path: str, finding, client: Client) -> Interpretation:
    """Ask the model for intent, then let the graph veto anything it invented."""
    ctx = build_context(path, finding)
    try:
        raw = client(SYSTEM_PROMPT, _render_user_prompt(ctx))
    except Exception as exc:  # noqa: BLE001 -- the audit must survive a model outage
        return Interpretation(ok=False, error=f"{type(exc).__name__}: {exc}")

    try:
        payload = json.loads(_strip_fence(raw))
    except (json.JSONDecodeError, TypeError) as exc:
        return Interpretation(ok=False, error=f"unparsable reply: {exc}")

    known = _known_cells(ctx)
    claimed = [str(c).upper().replace("$", "") for c in payload.get("cells_referenced", []) or []]
    rejected = [c for c in claimed if c not in known]

    interp = Interpretation(
        intent=_plain(str(payload.get("intent", ""))[:400]),
        deliberate=payload.get("deliberate"),
        explanation=_plain(str(payload.get("explanation", ""))[:800]),
        cells_referenced=[c for c in claimed if c in known],
        rejected_cells=rejected,
    )

    # A reference to a cell that was never in the context is a hallucination.
    # Sending an analyst to chase a cell that was never wrong is worse than silence.
    if rejected:
        interp.ok = False
        interp.error = f"referenced cells absent from the context: {', '.join(rejected)}"
    return interp


def _strip_fence(text: str) -> str:
    text = (text or "").strip()
    if text.startswith("```"):
        text = re.sub(r"^```[a-zA-Z]*\n", "", text)
        text = re.sub(r"\n```$", "", text)
    return text.strip()


def anthropic_client(
    model: str = DEFAULT_MODEL,
    max_tokens: int = 2000,
    effort: str = "medium",
) -> Callable[[str, str], str]:
    """Real client. Requires ANTHROPIC_API_KEY in the environment.

    Imported lazily so the deterministic arm never needs the SDK installed -- the
    whole point of the layering is that the product runs, and is measured, without
    a key or a network.

    `effort` is deliberately not `high`. The model's job here is small and bounded:
    read a dozen labelled cells and say what one of them was for. Recomputation has
    already established that it is wrong. Paying for deep reasoning on a question
    that narrow buys nothing, and the guard rejects invented references either way.
    """

    def call(system: str, user: str) -> str:
        import anthropic

        client = anthropic.Anthropic()
        resp = client.messages.create(
            model=model,
            max_tokens=max_tokens,
            system=system,
            thinking={"type": "adaptive"},
            output_config={"effort": effort},
            messages=[{"role": "user", "content": user}],
        )
        # A safety decline is a legitimate outcome, not an exception. Returning the
        # empty string lets `interpret` record it as an unparsable reply rather than
        # crashing an audit that has already produced valid deterministic findings.
        if resp.stop_reason == "refusal":
            return ""
        return "".join(b.text for b in resp.content if getattr(b, "type", "") == "text")

    return call
