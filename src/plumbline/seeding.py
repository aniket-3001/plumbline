"""Inject realistic errors into real workbooks, and record exactly what was injected.

Seeding realism decides whether the final number means anything. Invent errors and
the benchmark is fake-easy; a panel that sells rubric quality for a living will see
it immediately. So the taxonomy is Panko's, from the field-audit literature, not
one of our own devising:

    mechanical   typing and pointing slips -- the reference lands one row off
    logic        wrong formula for the intent -- AVERAGE where SUM belongs
    omission     something missing from the model -- a line left out of a total
    hardcoding   an input buried inside a formula, or a formula replaced by a
                 constant that is correct today and dead tomorrow

Design rules, each there for a reason:

  * **Seed into real Enron workbooks**, never synthetic ones. A detector tuned on
    tidy fixtures will not survive a real trading model.
  * **Only seed where a majority pattern exists.** An error is only findable if
    something establishes what "right" looked like. Seeding into a lone formula
    creates an unfindable case that silently punishes recall.
  * **Verify every seed actually changed the numbers** (except hardcode-dead, which
    is defined by *not* changing them today). A seed that alters nothing is not an
    error, it is noise in the ground truth.
  * **Record the pre-existing findings** in the untouched workbook. Real spreadsheets
    already contain real anomalies, so a detector hit that is not our seed is not
    automatically a false positive. Conflating those two would understate precision.
"""

from __future__ import annotations

import json
import random
import re
import sys
import warnings
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

_SCRIPTS = Path(__file__).resolve().parents[2] / "scripts"
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

from poc import A1_REF, native, normalise, split_ref  # noqa: E402

RANGE_REF = re.compile(r"(\$?[A-Z]{1,3}\$?[0-9]{1,7}):(\$?[A-Z]{1,3}\$?[0-9]{1,7})")


@dataclass
class Seed:
    """One injected error, and everything needed to score a detector against it."""

    seed_id: str
    panko_class: str
    sheet: str
    cell: str
    original_formula: str
    seeded_formula: str | object
    description: str
    detectable_by: list[str]
    baseline_value: object = None
    seeded_value: object = None
    value_changed: bool = True
    difficulty: str = "realistic"
    propagates_to: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return asdict(self)

    def classify_difficulty(self) -> str:
        """How much does this error announce itself?

        Reporting one blended accuracy number over both kinds would be
        misleading: a detector that only catches cells that broke loudly looks
        identical to one that catches silent corruption, and only the second is
        worth having. So the classes are scored separately.

          obvious    the cell now reads 0, blank or an Excel error. A human
                     scanning the sheet would very likely spot it.
          realistic  the cell still holds a plausible number. Nothing looks
                     wrong, and this is the error that survives review.
          silent     the value is unchanged today and only diverges later
                     (hardcoded formulas). The hardest class.
        """
        if self.panko_class == "hardcoding" and not self.value_changed:
            return "silent"
        v = self.seeded_value
        if v is None or v == 0 or (isinstance(v, str) and v.startswith("<error")):
            return "obvious"
        return "realistic"


def _majority_lines(
    formulas: dict[str, str], axis: str = "row"
) -> dict[int, list[tuple[str, int, str]]]:
    """Rows -- or columns -- where at least 3 formula cells agree on a single shape.

    These are the only places worth seeding: the agreement is what makes a
    deviation detectable in principle.

    The axis matters because a benchmark can only reward what it plants. Seeding
    exclusively along rows made the corpus structurally incapable of measuring
    column-wise detection: the column pass recovered **zero** of the 53 seeds, not
    because it is useless but because no error was ever placed where it could see
    one. A detector cannot be judged against a benchmark that never poses its
    question.
    """
    groups: dict[int, list[tuple[str, int, str]]] = defaultdict(list)
    for ref, text in formulas.items():
        try:
            row, col = split_ref(ref)
        except ValueError:
            continue
        line, position = (row, col) if axis == "row" else (col, row)
        groups[line].append((ref, position, text))

    out = {}
    for line, members in groups.items():
        if len(members) < 3:
            continue
        shapes = set()
        for ref, _pos, text in members:
            r, c = split_ref(ref)
            shapes.add(normalise(text, r, c))
        if len(shapes) == 1:  # unanimous -- the cleanest possible ground truth
            out[line] = members
    return out


def _majority_rows(formulas: dict[str, str]) -> dict[int, list[tuple[str, int, str]]]:
    """Backwards-compatible alias; the row axis is what v1..v5 measured."""
    return _majority_lines(formulas, "row")


def _shift_range_end(formula: str, delta: int) -> str | None:
    """Move the end of the first range by `delta` rows. Returns None if no range."""
    m = RANGE_REF.search(formula)
    if not m:
        return None
    end = m.group(2)
    em = re.fullmatch(r"(\$?)([A-Z]{1,3})(\$?)([0-9]{1,7})", end)
    if not em:
        return None
    new_row = int(em.group(4)) + delta
    if new_row < 1:
        return None
    new_end = f"{em.group(1)}{em.group(2)}{em.group(3)}{new_row}"
    return formula[: m.start(2)] + new_end + formula[m.end(2) :]


def _swap_function(formula: str, old: str, new: str) -> str | None:
    pattern = re.compile(rf"(?<![A-Za-z0-9_.]){old}\s*\(", re.I)
    if not pattern.search(formula):
        return None
    return pattern.sub(f"{new}(", formula, count=1)


def _shift_first_reference(formula: str, delta: int) -> str | None:
    """Move the first cell reference by `delta` rows -- the classic pointing slip.

    Most real formulas are plain references (`=D19`), not ranges, so a seeder that
    only knows how to shrink a SUM range finds nothing to do in a typical workbook.
    This is the more common mechanical error in the field: the formula was copied or
    dragged and landed one row off.
    """
    if RANGE_REF.search(formula):
        return None  # handled by _shift_range_end; keep the two classes distinct

    m = A1_REF.search(formula)
    if not m:
        return None
    col_abs, col_letters, row_abs, row_digits = m.groups()
    if row_abs:  # anchored on purpose; moving it would be a different error class
        return None
    new_row = int(row_digits) + delta
    if new_row < 1:
        return None
    replacement = f"{col_abs}{col_letters}{row_abs}{new_row}"
    return formula[: m.start()] + replacement + formula[m.end() :]


def plan_seeds(
    path: Path, rng: random.Random, max_seeds: int = 3, axes: tuple[str, ...] = ("row",)
) -> list[dict]:
    """Choose what to inject, without touching the file yet.

    `axes` defaults to row-only, which is what v1..v5 measured. Passing both plants
    errors that only a column-wise pass can see, so the corpus can finally ask that
    question -- but it also makes a different corpus, and therefore a new benchmark
    rather than the next rung of that ladder.
    """
    from poc import load_formulas

    sheets = load_formulas(str(path))
    candidates: list[dict] = []

    for sheet, formulas in sheets.items():
      for axis in axes:
        for line, members in _majority_lines(formulas, axis).items():
            # Seed the interior, never the first cell -- the majority must survive.
            for ref, col, text in members[1:]:
                up = text.upper()

                if RANGE_REF.search(text):
                    shifted = _shift_range_end(text, -1)
                    if shifted and shifted != text:
                        candidates.append(
                            dict(
                                panko_class="mechanical",
                                sheet=sheet,
                                cell=ref,
                                original_formula=text,
                                seeded_formula=shifted,
                                description="Range end pulled in by one row, dropping the last line.",
                                detectable_by=["pattern_break"],
                            )
                        )
                else:
                    pointed = _shift_first_reference(text, -1)
                    if pointed and pointed != text:
                        candidates.append(
                            dict(
                                panko_class="mechanical",
                                sheet=sheet,
                                cell=ref,
                                original_formula=text,
                                seeded_formula=pointed,
                                description=(
                                    "Reference lands one row above its neighbours' -- "
                                    "the formula was dragged or copied off by one."
                                ),
                                detectable_by=["pattern_break"],
                            )
                        )

                if "SUM(" in up:
                    swapped = _swap_function(text, "SUM", "AVERAGE")
                    if swapped:
                        candidates.append(
                            dict(
                                panko_class="logic",
                                sheet=sheet,
                                cell=ref,
                                original_formula=text,
                                seeded_formula=swapped,
                                description="SUM replaced by AVERAGE -- plausible formula, wrong intent.",
                                detectable_by=["pattern_break"],
                            )
                        )

                candidates.append(
                    dict(
                        panko_class="hardcoding",
                        sheet=sheet,
                        cell=ref,
                        original_formula=text,
                        seeded_formula=None,  # filled at injection time with the live value
                        description=(
                            "Formula replaced by its current value: correct today, "
                            "silently dead the moment an input moves."
                        ),
                        detectable_by=["dead_cell", "sensitivity_probe"],
                    )
                )

    rng.shuffle(candidates)

    # Three constraints, and the third one is not obvious.
    #
    # One seed per cell, and a per-class cap so no single Panko class dominates the
    # benchmark. Then: **at most one seed per row.**
    #
    # Detection is majority-vote within a row. Seed two cells of a three-formula row
    # and the two corrupted cells become the majority, so the one *correct* cell is
    # now the deviation. That happened on `john_zufferli__16801__marks.xlsx`, where
    # row 34 held I34/J34/K34 as `=AVERAGE(x10:x32)`; seeding J34 and K34 to `:31`
    # made the untouched I34 the outlier. The run scored two false negatives and one
    # false positive, and every one of those three judgements was the benchmark's
    # fault, not the detector's. A benchmark must not ask a tool to find something
    # that, by the benchmark's own definition, is no longer there.
    chosen: list[dict] = []
    used_cells: set[tuple[str, str]] = set()
    used_lines: set[tuple[str, str, int]] = set()
    per_class: dict[str, int] = defaultdict(int)
    cap = max(1, max_seeds // 2)
    for cand in candidates:
        key = (cand["sheet"], cand["cell"])
        row, col = _row_of(cand["cell"]), _col_of(cand["cell"])
        # One seed per row **and** per column. The row rule exists because two
        # errors in a three-formula row make the corrupted pair the majority and
        # the correct cell the outlier. Seeding along columns as well makes the
        # column the same hazard, and a corpus that plants both without this would
        # reintroduce the bug on the other axis.
        lines = {(cand["sheet"], "row", row), (cand["sheet"], "col", col)}
        if key in used_cells or lines & used_lines:
            continue
        if per_class[cand["panko_class"]] >= cap:
            continue
        chosen.append(cand)
        used_cells.add(key)
        used_lines |= lines
        per_class[cand["panko_class"]] += 1
        if len(chosen) >= max_seeds:
            break
    return chosen


def _row_of(ref: str) -> int:
    return int(re.sub(r"[^0-9]", "", ref) or 0)


def _col_of(ref: str) -> int:
    from openpyxl.utils import column_index_from_string

    letters = re.sub(r"[^A-Z]", "", ref.upper())
    return column_index_from_string(letters) if letters else 0


def apply_seeds(src: Path, dest: Path, plan: list[dict]) -> list[Seed]:
    """Write the seeded workbook and verify each seed actually did something."""
    from xlcalculator import Evaluator, ModelCompiler

    base_model = ModelCompiler().read_and_parse_archive(str(src))
    base_ev = Evaluator(base_model)

    def base_value(sheet: str, cell: str):
        try:
            return native(base_ev.evaluate(f"{sheet}!{cell}"))
        except Exception:  # noqa: BLE001
            return None

    wb = load_workbook(src)
    seeds: list[Seed] = []

    for i, cand in enumerate(plan):
        sheet, cell = cand["sheet"], cand["cell"]
        before = base_value(sheet, cell)
        seeded_formula = cand["seeded_formula"]
        if cand["panko_class"] == "hardcoding":
            if before is None:
                continue
            seeded_formula = before  # a constant equal to today's correct answer
        wb[sheet][cell] = seeded_formula
        seeds.append(
            Seed(
                seed_id=f"{src.stem}-{i}",
                panko_class=cand["panko_class"],
                sheet=sheet,
                cell=cell,
                original_formula=cand["original_formula"],
                seeded_formula=seeded_formula,
                description=cand["description"],
                detectable_by=cand["detectable_by"],
                baseline_value=before,
            )
        )

    if not seeds:
        return []

    wb.save(dest)

    # Verify: did each seed change the numbers? Hardcoding is the deliberate
    # exception -- it is defined by looking identical today.
    seeded_model = ModelCompiler().read_and_parse_archive(str(dest))
    seeded_ev = Evaluator(seeded_model)
    kept: list[Seed] = []
    for seed in seeds:
        try:
            after = native(seeded_ev.evaluate(f"{seed.sheet}!{seed.cell}"))
        except Exception:  # noqa: BLE001
            after = None
        seed.seeded_value = after
        seed.value_changed = after != seed.baseline_value
        if seed.panko_class == "hardcoding":
            # Must NOT have changed, or it is not the hard case we meant to build.
            if seed.value_changed:
                continue
        elif not seed.value_changed:
            # Changed nothing: not an error, just noise in the ground truth.
            continue
        seed.difficulty = seed.classify_difficulty()
        kept.append(seed)

    if len(kept) != len(seeds):
        # Rewrite with only the verified seeds so the file matches the manifest.
        wb2 = load_workbook(src)
        for seed in kept:
            wb2[seed.sheet][seed.cell] = seed.seeded_formula
        wb2.save(dest)

    return kept


def pre_existing_findings(
    path: Path, *, min_peers: int | None = None, contiguous: bool = True
) -> list[str]:
    """Cells the *original* workbook already flags, before any seed is injected.

    Scoring excludes these rather than counting them against precision: they are
    not our errors, we have no ground truth for them, and Enron's spreadsheets are
    full of them -- 144 across 21 workbooks in the first run.

    This must run **every detector the audit runs**, on the **unseeded** file, and
    **at the same settings**. All three halves have bitten. `min_peers` is passed
    through for the third: raise the audit's sensitivity without raising the
    exclusion list's, and every extra pre-existing cell the audit now finds is
    charged to it as a false positive, so a detector improvement reads as a
    precision collapse.

    Getting the first half wrong is expensive and silent. An earlier version ran
    only the pattern-break detector, so every pre-existing *dead cell* fell through
    as a false positive: on `scott_neal__38672` six typed constants sitting in
    `=Z41+1` counter rows -- real hardcoding, present in the file Enron shipped --
    were charged to Plumbline. Eleven of the twelve false positives in the first
    run were this one bug.

    Getting the second half wrong would be worse and would not look like a bug at
    all: run this on the seeded copy and it excludes the seeds themselves, and
    recall becomes meaningless while every number still looks plausible.
    """
    from plumbline.audit import (
        MIN_ROW_PEERS,
        detect_dead_cells,
        detect_pattern_breaks,
        screen_dead_cells,
    )
    from poc import load_formulas

    min_peers = MIN_ROW_PEERS if min_peers is None else min_peers

    try:
        sheets = load_formulas(str(path))
    except Exception:  # noqa: BLE001
        return []

    refs: list[str] = []
    dead: list = []
    for sheet, formulas in sheets.items():
        try:
            refs.extend(f"{sheet}!{f.cell}" for f in detect_pattern_breaks(sheet, formulas))
            dead.extend(detect_dead_cells(str(path), sheet, formulas, min_peers=min_peers,
                                          contiguous=contiguous))
        except Exception:  # noqa: BLE001
            continue

    # Screened the same way the audit screens, so the two agree cell for cell.
    try:
        refs.extend(f"{f.sheet}!{f.cell}" for f in screen_dead_cells(str(path), dead))
    except Exception:  # noqa: BLE001
        pass
    return sorted(set(refs))


def seed_workbook(src: Path, dest_dir: Path, rng: random.Random, max_seeds: int = 3) -> dict | None:
    """Seed one workbook and return its ground-truth manifest."""
    from poc import detect_row_pattern_breaks, load_formulas

    plan = plan_seeds(src, rng, max_seeds=max_seeds)
    if not plan:
        return None

    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / src.name
    seeds = apply_seeds(src, dest, plan)
    if not seeds:
        if dest.exists():
            dest.unlink()
        return None

    # What does the untouched workbook already flag? Those are not our seeds, and
    # counting them as false positives would understate precision.
    pre_existing = pre_existing_findings(src)

    manifest = {
        "workbook": src.name,
        "source": str(src),
        "seeded": dest.name,
        "seed_count": len(seeds),
        "pre_existing_findings": pre_existing,
        "seeds": [s.to_dict() for s in seeds],
    }
    (dest_dir / f"{src.stem}.truth.json").write_text(
        json.dumps(manifest, indent=2, default=str), encoding="utf-8"
    )
    return manifest
