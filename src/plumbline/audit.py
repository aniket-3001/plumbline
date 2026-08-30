"""The deterministic audit pass: detect, then prove.

This is Layer 1 and Layer 3 of the architecture with no model in the loop. It is
both a working product on its own and the baseline the model layers must beat --
running it standalone is what makes the ablation in the changelog meaningful.

Contract: a finding leaves this module only if a recomputation demonstrates it.
Everything else is reported as unproved and, under the strict contract, dropped.
"""

from __future__ import annotations

import os
import re
import sys
import tempfile
import warnings
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path

warnings.filterwarnings("ignore")

_SCRIPTS = Path(__file__).resolve().parents[2] / "scripts"
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

from openpyxl.utils import column_index_from_string  # noqa: E402

from poc import native, normalise, rebase, split_ref  # noqa: E402


@dataclass
class Finding:
    sheet: str
    cell: str
    detector: str
    panko_class: str
    actual: str
    expected: str
    reason: str
    proved: bool = False
    proof: str = ""
    baseline_value: object = None
    repaired_value: object = None
    #: How much the reported figure moves if this cell is corrected **today**.
    #: For a dead cell that is zero by definition -- being correct today is what
    #: makes it dead -- so `delta` stays None there and the probe's effect goes in
    #: `probe_response`. Conflating the two put the probe's perturbation size into
    #: the report's "largest single correction" headline, which was simply false.
    delta: object = None
    probe_response: object = None
    #: Which orientation found this. Two detectors can flag the same cell from
    #: different directions, and the report says which so a reader can check it.
    axis: str = "row"
    impacted: list = field(default_factory=list)

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class AuditReport:
    workbook: str
    findings: list[Finding] = field(default_factory=list)
    formula_cells: int = 0
    skipped: str | None = None
    dead_candidates: int = 0
    dead_screened_out: int = 0
    detected: int = 0        # before any cap
    proof_truncated: bool = False
    proof_deferred: int = 0  # detected and reported, but not proved: budget ran out

    @property
    def proved(self) -> list[Finding]:
        """Proved findings, largest money impact first.

        Detection order is an artifact of how the sheets happened to be walked, and
        it is not the order anyone reads in. The user is signing off a board pack:
        a correction worth three million belongs above one worth thirty, because if
        they only get through half the list, that is the half that has to matter.

        Dead cells have no delta -- repairing one changes nothing today, which is
        the whole reason they need a sensitivity probe -- so they sort after the
        cells whose impact is a number, and by address among themselves so the order
        is at least stable between runs.
        """
        def impact(f: Finding):
            delta = f.delta if isinstance(f.delta, (int, float)) else None
            return (0 if delta is not None else 1,
                    -abs(delta) if delta is not None else 0,
                    f.sheet, f.cell)

        return sorted((f for f in self.findings if f.proved), key=impact)

    def to_dict(self) -> dict:
        return {
            "workbook": self.workbook,
            "formula_cells": self.formula_cells,
            "skipped": self.skipped,
            "dead_candidates": self.dead_candidates,
            "dead_screened_out": self.dead_screened_out,
            "detected": self.detected,
            "proof_truncated": self.proof_truncated,
            "proof_deferred": self.proof_deferred,
            "findings": [f.to_dict() for f in self.findings],
            "counts": {
                "total": len(self.findings),
                "proved": len(self.proved),
                "by_detector": dict(Counter(f.detector for f in self.findings)),
            },
        }


def _load(path: str):
    from xlcalculator import Evaluator, ModelCompiler

    model = ModelCompiler().read_and_parse_archive(path)
    return model, Evaluator(model)


def _formulas_by_sheet(model) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for addr, cell in model.cells.items():
        text = getattr(cell.formula, "formula", None)
        if not text:
            continue
        sheet, _, ref = addr.rpartition("!")
        out.setdefault(sheet, {})[ref] = text
    return out


#: The two orientations a spreadsheet pattern can run in.
#:
#: Everything here was row-only until v6, and the report said so in its blind-spots
#: section. That was half the surface: a model has line items down the rows and
#: periods across the columns, so a growth-rate column or a running total is a
#: pattern the row-wise pass cannot see at all. The detectors are identical apart
#: from which coordinate groups and which varies, so the axis is a parameter rather
#: than a second copy of the logic that would drift from the first.
AXES = ("row", "col")


def _grouped(formulas: dict[str, str], axis: str):
    """(line, [(ref, position, text)]) for each row -- or each column.

    `position` is the coordinate that varies along the line, which is what both the
    majority vote and the nearest-peer search need; `line` is the one that is fixed.
    """
    groups: dict[int, list[tuple[str, int, str]]] = defaultdict(list)
    for ref, text in formulas.items():
        try:
            row, col = split_ref(ref)
        except ValueError:
            continue
        line, position = (row, col) if axis == "row" else (col, row)
        groups[line].append((ref, position, text))
    return groups


def detect_pattern_breaks(
    sheet: str, formulas: dict[str, str], *, axis: str = "row"
) -> list[Finding]:
    """A formula cell whose shape disagrees with its neighbours along one axis."""
    findings: list[Finding] = []
    for line, members in _grouped(formulas, axis).items():
        if len(members) < 3:
            continue

        # Drop aggregates *before* the vote, not after. A total is a deviant by any
        # measure, so leaving it in means a line holding both a total and a real
        # error has two deviants, fails the "exactly one" rule, and reports nothing.
        # Filtering first restores the error to being the only deviant.
        positions = [p for _, p, _ in members]
        members = [
            (r, p, t) for r, p, t in members
            if not _aggregates_its_peers(t, axis, [q for q in positions if q != p])
        ]
        if len(members) < 3:
            continue

        shapes = {}
        for ref, _pos, text in members:
            r, c = split_ref(ref)
            shapes[ref] = normalise(text, r, c)
        counts = Counter(shapes.values())
        majority, n = counts.most_common(1)[0]
        # Require a clear majority: exactly one deviant, everyone else agreeing.
        if n != len(members) - 1:
            continue
        conformers = [(r, p, t) for r, p, t in members if shapes[r] == majority]
        for ref, pos, text in members:
            if shapes[ref] == majority:
                continue
            # Translate from the NEAREST conforming peer. Rebasing shifts every
            # relative reference by the distance, so a distant peer can push
            # references off the edge of the sheet entirely.
            twin_ref, _twin_pos, twin_text = min(conformers, key=lambda m: abs(m[1] - pos))
            twin_row, twin_col = split_ref(twin_ref)
            row, col = split_ref(ref)
            expected = rebase(twin_text, twin_row, twin_col, row, col)
            if expected is None or expected == text:
                continue  # no usable expectation; proposing nothing beats proposing nonsense
            where = f"row {line}" if axis == "row" else f"column {_col_name(line)}"
            findings.append(
                Finding(
                    sheet=sheet,
                    cell=ref,
                    detector="pattern_break",
                    panko_class="mechanical/logic",
                    actual=text,
                    expected=expected,
                    axis=axis,
                    reason=(
                        f"{n} of {len(members)} formula cells in {where} share one shape; "
                        f"{ref} does not."
                    ),
                )
            )
    return findings


def _dedupe(findings: list[Finding]) -> list[Finding]:
    """One finding per cell, preferring the row reading when both axes agree."""
    best: dict[tuple[str, str], Finding] = {}
    for f in findings:
        key = (f.sheet, f.cell)
        if key not in best or (best[key].axis != "row" and f.axis == "row"):
            best[key] = f
    return list(best.values())


def _where(axis: str, line: int) -> str:
    return f"row {line}" if axis == "row" else f"column {_col_name(line)}"


RANGE_RE = re.compile(r"([A-Z]{1,3})([1-9][0-9]{0,6}):([A-Z]{1,3})([1-9][0-9]{0,6})")


def _aggregates_its_peers(text: str, axis: str, peer_positions: list[int]) -> bool:
    """Is this cell the total of the very cells it is being compared against?

    The single most common shape in any spreadsheet is a run of values with a sum
    at the end of it. Along that run the sum is the only cell with a different
    formula, so a naive majority vote calls the **total** the error -- and does it
    everywhere, because totals are everywhere.

    Measured on the corpus before this guard existed, the column pass produced 235
    findings the row pass missed, and on `darrell_schoolcraft__7407` it went from 1
    to 124: rows 7-30 are hourly values and row 31 is `=SUM(E7:E30)`, flagged as a
    break in all four columns. Those were not errors, and the benchmark could not
    have told us -- they are all pre-existing, so they land in the excluded bucket
    and precision would have looked untouched. Same shape as the `min_peers`
    ablation, caught this time before it shipped.

    A cell that spans its own peers is aggregating them, not failing to copy them.
    """
    if not peer_positions:
        return False
    lo, hi = min(peer_positions), max(peer_positions)
    for c1, r1, c2, r2 in RANGE_RE.findall(text or ""):
        if axis == "row":
            start, end = column_index_from_string(c1), column_index_from_string(c2)
        else:
            start, end = int(r1), int(r2)
        covered = [p for p in peer_positions if start <= p <= end]
        # Half is enough: a subtotal often spans a block of the run rather than
        # all of it, and a genuine off-by-one copy never spans its neighbours.
        if len(covered) >= max(2, len(peer_positions) // 2) and start <= hi and end >= lo:
            return True
    return False


def _col_name(index: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(index)


#: How many formula peers a row needs before a typed constant in it means anything.
#:
#: Measured, not argued -- see `Docs/MIN_PEERS_ABLATION.md`. Lowering this to 2
#: takes recall from 0.868 to 0.981 at unchanged precision, and that reading is
#: wrong: the 29 extra findings are all *pre-existing*, and pre-existing findings
#: are excluded from scoring, so the benchmark shows the benefit and is structurally
#: blind to the cost. Reading all 29 by hand, roughly ten are ordinary data flagged
#: wrongly and sixteen are unverifiable zeros.
#:
#: Lowering this to 2 was measured as a clear win *only once `_peers_in_block` existed*.
#: Without it, 2 took recall 0.868 -> 0.981 while quietly adding 29 unverified
#: pre-existing findings, roughly ten of which were hand-labelled as ordinary data.
#: With it, 2 gives recall 0.924 and the unverified population goes *down*, 368 -> 362.
#:
#:      no contiguity, 3   recall 0.868  F1 0.929  pre-existing 368   <- was shipped
#:      no contiguity, 2   recall 0.981  F1 0.991  pre-existing 397
#:      contiguity,    3   recall 0.811  F1 0.896  pre-existing 360
#:      contiguity,    2   recall 0.924  F1 0.961  pre-existing 362   <- ships
#:
#: Precision is 1.000 in all four, so the threshold was never the precision knob it
#: looked like; block membership was. See `Docs/MIN_PEERS_ABLATION.md`.
MIN_ROW_PEERS = 2


def _peers_in_block(col: int, cols: list[int]) -> list[int]:
    """The peer columns that belong to the same block of the row as `col`.

    "Block" means a run of cells laid out on one regular stride. Financial sheets
    lay out either densely (`C D E F`) or on a spacer rhythm (`C _ E _ G`), and both
    are one coherent block; what separates two blocks is a change of rhythm or a
    stretch of unrelated content.

    So a peer counts when the candidate is **on the block's stride and inside its
    span**. Two failures drove that wording, and both were real sheets:

      Report!Y50   C E G I K M O Q S U W _ AA -- peers every second column, so
                   strict adjacency saw Y50's blank neighbours as a boundary and
                   threw away a seeded error. Stride 2 accepts it.
      MANUAL!D28   B _ _ _ _ G -- two counter columns with a name and an id
                   between. D is inside the span but off the stride of 5, so it is
                   still correctly rejected as the data column it is.
      Options!F16  B _ D -- stride 2, and F continues that rhythm, so this returns
                   two peers and the block rule does *not* reject it. Only
                   `min_peers = 3` does. Recorded here because it is the honest
                   limit of the idea: a data column that happens to sit on the same
                   rhythm as a nearby formula block is indistinguishable by layout
                   alone, and would need the labels to settle.

    A candidate whose block contains no other formula at all is unreachable this
    way, by construction. `Floor Plan!Y41` is the case: its block is `X41 Y41 Z41`
    with plain numbers either side, so nothing in the row can vouch for it.
    """
    if not cols:
        return []
    ordered = sorted(set(cols) | {col})
    i = ordered.index(col)

    # The stride is the rhythm the neighbours actually keep. With one neighbour there
    # is no rhythm to read, so fall back to the distance to it.
    gaps = [b - a for a, b in zip(ordered, ordered[1:])]
    stride = min(gaps) if gaps else 1

    block = [col]
    for direction in (-1, 1):
        j, expected = i + direction, col + direction * stride
        while 0 <= j < len(ordered) and ordered[j] == expected:
            block.append(ordered[j])
            j += direction
            expected += direction * stride
    return [c for c in block if c != col]


def detect_dead_cells(
    path: str,
    sheet: str,
    formulas: dict[str, str],
    *,
    min_peers: int = MIN_ROW_PEERS,
    contiguous: bool = True,
    axis: str = "row",
) -> list[Finding]:
    """A typed constant sitting where every neighbour holds a formula."""
    from openpyxl import load_workbook

    grouped = _grouped(formulas, axis)

    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb[sheet] if sheet in wb.sheetnames else None
        if ws is None:
            return []
        constants: dict[str, object] = {}
        for r in ws.iter_rows():
            for cell in r:
                v = cell.value
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    constants[cell.coordinate] = v
    finally:
        wb.close()

    findings: list[Finding] = []
    for ref, value in constants.items():
        try:
            row, col = split_ref(ref)
        except ValueError:
            continue
        line, pos = (row, col) if axis == "row" else (col, row)
        peers = grouped.get(line, [])
        if len(peers) < min_peers:
            continue
        shapes = set()
        for pref, _p, t in peers:
            pr, pc = split_ref(pref)
            shapes.add(normalise(t, pr, pc))
        if len(shapes) != 1:
            continue  # peers disagree among themselves; no clean expectation

        if contiguous:
            near = set(_peers_in_block(pos, [p for _, p, _ in peers]))
            if len(near) < min_peers:
                continue  # peers exist, but in another block of this line
            peers = [(r, p, t) for r, p, t in peers if p in near]

        twin_ref, _twin_pos, twin_text = min(peers, key=lambda m: abs(m[1] - pos))
        twin_row, twin_col = split_ref(twin_ref)
        expected = rebase(twin_text, twin_row, twin_col, row, col)
        if expected is None:
            continue  # translation would land off-sheet; no expectation to offer
        findings.append(
            Finding(
                sheet=sheet,
                cell=ref,
                detector="dead_cell",
                panko_class="hardcoding",
                actual=str(value),
                expected=expected,
                axis=axis,
                reason=(
                    f"{len(peers)} adjacent cells in {_where(axis, line)} share one "
                    f"formula shape; {ref} sits among them as a typed constant."
                    if contiguous else
                    f"all {len(peers)} other cells in {_where(axis, line)} share one "
                    f"formula shape; {ref} is a typed constant."
                ),
            )
        )
    return findings


def _close(a, b, rel: float = 1e-6) -> bool:
    """Numeric equality with a tolerance, so float noise is not read as a difference."""
    if a is None or b is None:
        return False
    if isinstance(a, bool) or isinstance(b, bool):
        return a == b
    try:
        fa, fb = float(a), float(b)
    except (TypeError, ValueError):
        return str(a) == str(b)
    return abs(fa - fb) <= rel * max(1.0, abs(fa), abs(fb))


def screen_dead_cells(path: str, candidates: list[Finding]) -> list[Finding]:
    """Keep only constants that look like the frozen output of their row's formula.

    Without this, the detector drowns in false positives on real sheets. Enron
    workbooks are full of rows like `A7=data, B7=30468, C7==+B7, D7==+C7, E7=10000`
    -- a carry-forward chain where some cells are genuine typed inputs sitting
    among formulas. Flagging every input as a dead formula produced 40 false
    positives on a single workbook.

    The discriminator applies Plumbline's own thesis: compute what the expected
    formula would actually yield in that position. If it equals the constant, the
    cell really does look like a formula someone froze. If it differs, the constant
    is data that merely happens to live in a formula row.

    Each expected formula is evaluated in a **scratch cell**, not in the candidate's
    own position. Writing them all back in place was the obvious approach and it is
    wrong: one candidate's expected formula frequently reads another candidate, so
    the replacements cascade and every value comes out contaminated. Measured on a
    real workbook, in-place screening discarded all 41 candidates including the
    known seeded error.

    A1 references are literal text, so a formula means the same thing wherever it
    sits on its own sheet. Parking each one in a far-right scratch column leaves
    every original cell untouched, needs a single extra parse for the whole
    workbook, and cannot cascade.

    Deliberate limitation: a hardcode whose value has already drifted away from its
    inputs is dropped. That case is far more often intentional data than a frozen
    formula, and the silent-time-bomb case we care about -- correct today, wrong
    tomorrow -- is precisely the one where constant and formula still agree.
    """
    if not candidates:
        return []

    from openpyxl.utils import get_column_letter

    SCRATCH_COL = 16000  # far right of any real content, well inside Excel's XFD limit
    tmp = os.path.join(tempfile.gettempdir(), f"plumbline_screen_{os.getpid()}.xlsx")

    try:
        from openpyxl import load_workbook

        wb = load_workbook(path)
        scratch: dict[int, Finding] = {}
        per_sheet_row: dict[str, int] = defaultdict(int)
        for i, f in enumerate(candidates):
            if f.sheet not in wb.sheetnames:
                continue
            per_sheet_row[f.sheet] += 1
            row = per_sheet_row[f.sheet]
            col_offset = row // 1_000_000  # stay inside Excel's row limit
            addr = f"{get_column_letter(SCRATCH_COL + col_offset)}{(row % 1_000_000) + 1}"
            wb[f.sheet][addr] = f.expected
            scratch[i] = (f, f"{f.sheet}!{addr}")
        if not scratch:
            return candidates
        wb.save(tmp)

        _, ev = _load(tmp)
        kept = []
        for _, (f, addr) in scratch.items():
            try:
                would_be = native(ev.evaluate(addr))
            except Exception:  # noqa: BLE001
                continue
            if _close(would_be, _as_number(f.actual)):
                kept.append(f)
        return kept
    except Exception:  # noqa: BLE001 -- screening must never abort an audit
        return candidates
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)


def _as_number(text: str):
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _write_variant(path: str, edits: dict[tuple[str, str], object], dest: str) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    for (sheet, cell), value in edits.items():
        wb[sheet][cell] = value
    wb.save(dest)


def prove(path: str, findings: list[Finding], watch_limit: int = 250) -> list[Finding]:
    """Recompute with each proposed repair applied; the delta is the proof.

    Pattern breaks are proved directly: apply the repair, show the numbers move.
    Dead cells need the indirect route -- the constant is already correct today, so
    repairing it changes nothing. Instead perturb an input it should depend on and
    show it fails to respond while the repaired version does.
    """
    if not findings:
        return findings

    _, base_ev = _load(path)

    def value_of(sheet: str, cell: str):
        try:
            return native(base_ev.evaluate(f"{sheet}!{cell}"))
        except Exception:  # noqa: BLE001
            return None

    tmp = os.path.join(tempfile.gettempdir(), f"plumbline_audit_{os.getpid()}.xlsx")
    tmp2 = os.path.join(tempfile.gettempdir(), f"plumbline_audit2_{os.getpid()}.xlsx")

    try:
        for finding in findings:
            key = (finding.sheet, finding.cell)
            finding.baseline_value = value_of(*key)

            if finding.detector == "pattern_break":
                _write_variant(path, {key: finding.expected}, tmp)
                try:
                    _, ev = _load(tmp)
                    finding.repaired_value = native(ev.evaluate(f"{finding.sheet}!{finding.cell}"))
                except Exception as exc:  # noqa: BLE001
                    finding.proof = f"recomputation failed: {type(exc).__name__}"
                    continue
                try:
                    finding.delta = finding.repaired_value - finding.baseline_value
                except TypeError:
                    finding.delta = None
                if finding.delta not in (None, 0):
                    finding.proved = True
                    finding.proof = (
                        f"{finding.cell}: {finding.baseline_value} -> "
                        f"{finding.repaired_value} ({finding.delta:+})"
                    )
                else:
                    finding.proof = "repair changes nothing; not reported"

            elif finding.detector == "dead_cell":
                # Find an input the expected formula depends on, and nudge it.
                from poc import A1_REF

                refs = [m.group(0).replace("$", "") for m in A1_REF.finditer(finding.expected)]
                if not refs:
                    finding.proof = "no inputs to perturb"
                    continue
                target = refs[0]
                original = value_of(finding.sheet, target)
                if not isinstance(original, (int, float)) or isinstance(original, bool):
                    finding.proof = f"input {target} is not numeric"
                    continue

                perturbed = original + 1000
                # Arm A: perturb the input, leave the suspect cell as it is.
                _write_variant(path, {(finding.sheet, target): perturbed}, tmp)
                # Arm B: perturb the input AND restore the expected formula.
                _write_variant(
                    path,
                    {(finding.sheet, target): perturbed, key: finding.expected},
                    tmp2,
                )
                try:
                    _, ev_a = _load(tmp)
                    _, ev_b = _load(tmp2)
                    as_is = native(ev_a.evaluate(f"{finding.sheet}!{finding.cell}"))
                    as_formula = native(ev_b.evaluate(f"{finding.sheet}!{finding.cell}"))
                except Exception as exc:  # noqa: BLE001
                    finding.proof = f"recomputation failed: {type(exc).__name__}"
                    continue

                finding.repaired_value = as_formula
                unresponsive = as_is == finding.baseline_value
                control_responded = as_formula != finding.baseline_value
                if unresponsive and control_responded:
                    finding.proved = True
                    try:
                        # The probe's effect, not a correction: it is the size of the
                        # perturbation this cell failed to follow. Never a money figure.
                        finding.probe_response = as_formula - as_is
                    except TypeError:
                        finding.probe_response = None
                    finding.proof = (
                        f"set {target} {original} -> {perturbed}: "
                        f"{finding.cell} as-is {finding.baseline_value} -> {as_is} (no response); "
                        f"as formula -> {as_formula} (responds)"
                    )
                else:
                    finding.proof = (
                        "cell responded to its inputs, or the control did not; not reported"
                    )
    finally:
        for f in (tmp, tmp2):
            if os.path.exists(f):
                os.remove(f)

    return findings


def audit(
    path: str | Path,
    *,
    check_determinism: bool = True,
    max_proofs: int = 0,
    min_peers: int = MIN_ROW_PEERS,
    contiguous: bool = True,
    axes: tuple[str, ...] = AXES,
) -> AuditReport:
    """Full deterministic audit of one workbook."""
    from plumbline.determinism import check, find_volatile

    path = str(path)
    report = AuditReport(workbook=Path(path).name)

    if check_determinism:
        vol = find_volatile(path)
        if vol.is_volatile:
            report.skipped = f"volatile: {vol.summary()}"
            return report
        det = check(path, limit=150)
        if not det.stable:
            report.skipped = f"nondeterministic: {det.summary()}"
            return report

    try:
        model, _ = _load(path)
    except Exception as exc:  # noqa: BLE001
        report.skipped = f"parse failed: {type(exc).__name__}"
        return report

    sheets = _formulas_by_sheet(model)
    report.formula_cells = sum(len(v) for v in sheets.values())

    findings: list[Finding] = []
    dead_candidates: list[Finding] = []
    for sheet, formulas in sheets.items():
        for axis in axes:
            findings.extend(detect_pattern_breaks(sheet, formulas, axis=axis))
            dead_candidates.extend(
                detect_dead_cells(path, sheet, formulas, min_peers=min_peers,
                                  contiguous=contiguous, axis=axis)
            )

    # A genuinely wrong cell is often wrong in both directions at once, and the same
    # address reported twice is a bug in the report, not two findings. Keep the row
    # reading when both agree -- it is the one every earlier measurement used, so the
    # v1..v5 numbers stay comparable -- and keep whichever arrived first otherwise.
    findings = _dedupe(findings)
    dead_candidates = _dedupe(dead_candidates)

    report.dead_candidates = len(dead_candidates)
    screened = screen_dead_cells(path, dead_candidates)
    report.dead_screened_out = len(dead_candidates) - len(screened)
    findings.extend(screened)
    report.detected = len(findings)

    # Proving costs one full workbook re-parse per finding, so a large sheet with
    # many candidates can run for many minutes. The budget therefore caps *proofs*,
    # never detection: every detected cell is still reported, and the ones past the
    # budget are carried through explicitly unproved rather than discarded.
    #
    # An earlier version sliced `findings` itself. That silently deleted detections,
    # and on four corpus workbooks the deleted findings included the seeded errors --
    # so the cap was scored as a detector failure. A budget must never be able to
    # change what the tool claims to have looked at.
    to_prove, deferred = findings, []
    if max_proofs and len(findings) > max_proofs:
        to_prove, deferred = findings[:max_proofs], findings[max_proofs:]
        report.proof_truncated = True
        report.proof_deferred = len(deferred)
        for f in deferred:
            f.proof = "not attempted: proof budget exhausted"

    report.findings = prove(path, to_prove) + deferred
    return report
