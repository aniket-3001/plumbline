"""Proof of concept: detect a pattern break, then PROVE it by recomputation.

This is the whole Plumbline thesis in miniature, with no model in the loop yet:

  1. Parse the workbook into a formula model.
  2. Normalise each formula to relative (R1C1-style) offsets so that cells doing
     "the same thing" collapse to the same string.
  3. Flag any cell whose normalised formula disagrees with its row neighbours.
  4. Propose the majority formula as the repair.
  5. Recompute the workbook with the repair applied and report the delta --
     which is the proof. No delta, no finding.

Usage:  python scripts/poc.py [workbook.xlsx]
"""

from __future__ import annotations

import os
import re
import sys
import tempfile
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from xlcalculator import Evaluator, ModelCompiler

# Matches an A1-style reference, optionally absolute, not preceded by a letter
# (so it will not chew the tail of a function name such as LOG10).
A1_REF = re.compile(r"(?<![A-Za-z0-9_])(\$?)([A-Z]{1,3})(\$?)([1-9][0-9]{0,6})(?![0-9(])")


def native(value):
    """xlcalculator returns its own numeric wrappers; unwrap to plain Python."""
    if isinstance(value, bool) or value is None:
        return value
    try:
        f = float(value)
    except (TypeError, ValueError):
        return value
    return int(f) if f.is_integer() else f


@dataclass
class Finding:
    sheet: str
    cell: str
    actual: str
    expected: str
    reason: str
    baseline_value: object = None
    repaired_value: object = None
    delta: object = None
    impacted: tuple = ()

    @property
    def proved(self) -> bool:
        """A finding is only real if repairing it actually changes the numbers."""
        return self.delta not in (None, 0, 0.0)


def normalise(formula: str, row: int, col: int) -> str:
    """Rewrite absolute A1 references as offsets relative to the host cell.

    =SUM(B8:B10) in B11  ->  =SUM(R[-3]C[0]:R[-1]C[0])
    =SUM(C8:C10) in C11  ->  =SUM(R[-3]C[0]:R[-1]C[0])   (identical: same pattern)
    =SUM(C8:C9)  in C11  ->  =SUM(R[-3]C[0]:R[-2]C[0])   (different: pattern break)
    """

    def repl(m: re.Match) -> str:
        col_abs, col_letters, row_abs, row_digits = m.groups()
        ref_col = column_index_from_string(col_letters)
        ref_row = int(row_digits)
        col_part = f"C{ref_col}" if col_abs else f"C[{ref_col - col:+d}]".replace("+", "")
        row_part = f"R{ref_row}" if row_abs else f"R[{ref_row - row:+d}]".replace("+", "")
        return f"{row_part}{col_part}"

    return A1_REF.sub(repl, formula.upper())


def load_formulas(path: str) -> dict[str, dict[str, str]]:
    """{sheet: {A1_address: formula_text}} for every formula cell."""
    model = ModelCompiler().read_and_parse_archive(path)
    out: dict[str, dict[str, str]] = {}
    for addr, cell in model.cells.items():
        text = getattr(cell.formula, "formula", None)
        if not text:
            continue
        sheet, _, ref = addr.rpartition("!")
        out.setdefault(sheet, {})[ref] = text
    return out


def split_ref(ref: str) -> tuple[int, int]:
    m = re.fullmatch(r"([A-Z]{1,3})([0-9]+)", ref)
    if not m:
        raise ValueError(f"unparsable reference {ref!r}")
    return int(m.group(2)), column_index_from_string(m.group(1))


def detect_row_pattern_breaks(formulas: dict[str, str]) -> list[Finding]:
    """Group formula cells by row; flag minority patterns against the majority."""
    by_row: dict[int, list[tuple[str, int, str]]] = {}
    for ref, text in formulas.items():
        row, col = split_ref(ref)
        by_row.setdefault(row, []).append((ref, col, text))

    findings: list[Finding] = []
    for row, members in by_row.items():
        if len(members) < 3:  # too few peers to call anything a pattern
            continue
        shapes = {ref: normalise(text, row, col) for ref, col, text in members}
        counts = Counter(shapes.values())
        majority, majority_n = counts.most_common(1)[0]
        if majority_n < len(members) - 1 or majority_n == len(members):
            continue  # no clear majority, or everyone agrees
        for ref, col, text in members:
            if shapes[ref] == majority:
                continue
            conformers = [(r, c, t) for r, c, t in members if shapes[r] == majority]
            twin_ref, twin_col, twin = min(conformers, key=lambda m: abs(m[1] - col))
            twin_row, _ = split_ref(twin_ref)
            expected = rebase(twin, twin_row, twin_col, row, col)
            if expected is None or expected == text:
                continue
            findings.append(
                Finding(
                    sheet="",
                    cell=ref,
                    actual=text,
                    expected=expected,
                    reason=(
                        f"{majority_n} of {len(members)} formula cells in row {row} share one "
                        f"shape; {ref} does not."
                    ),
                )
            )
    return findings


MAX_COL = 16384  # Excel's last column, XFD


def rebase(formula: str, from_row: int, from_col: int, to_row: int, to_col: int) -> str | None:
    """Translate a formula written for one cell into the equivalent for another.

    Returns None when the translation would land outside the sheet. That happens
    for real: if the reference peer sits far to the right of the target cell, every
    relative column shifts left by that distance and can run off the left edge.
    Callers must treat None as "no expectation available" and skip the finding --
    a translated formula that points nowhere is not a repair worth proposing.
    """
    failed = False

    def repl(m: re.Match) -> str:
        nonlocal failed
        col_abs, col_letters, row_abs, row_digits = m.groups()
        ref_col = column_index_from_string(col_letters)
        ref_row = int(row_digits)
        new_col = ref_col if col_abs else ref_col + (to_col - from_col)
        new_row = ref_row if row_abs else ref_row + (to_row - from_row)
        if not (1 <= new_col <= MAX_COL) or new_row < 1:
            failed = True
            return m.group(0)
        return f"{col_abs}{get_column_letter(new_col)}{row_abs}{new_row}"

    out = A1_REF.sub(repl, formula.upper())
    return None if failed else out


def prove(path: str, sheet: str, finding: Finding, watch: list[str]) -> Finding:
    """Recompute the workbook with the repair applied. The delta is the proof.

    Note: we write a patched copy and re-parse it. The in-memory shortcuts
    (Evaluator.set_cell_value, or swapping the XLFormula and calling build_code)
    both return silently wrong numbers -- see Docs/DESIGN.md.
    """
    model = ModelCompiler().read_and_parse_archive(path)
    ev = Evaluator(model)
    finding.baseline_value = native(ev.evaluate(f"{sheet}!{finding.cell}"))
    before = {ref: native(ev.evaluate(f"{sheet}!{ref}")) for ref in watch}

    wb = load_workbook(path)
    wb[sheet][finding.cell] = finding.expected
    tmp = os.path.join(tempfile.gettempdir(), f"plumbline_cf_{os.getpid()}.xlsx")
    try:
        wb.save(tmp)
        model2 = ModelCompiler().read_and_parse_archive(tmp)
        ev2 = Evaluator(model2)
        finding.repaired_value = native(ev2.evaluate(f"{sheet}!{finding.cell}"))
        after = {ref: native(ev2.evaluate(f"{sheet}!{ref}")) for ref in watch}
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)

    try:
        finding.delta = finding.repaired_value - finding.baseline_value
    except TypeError:
        finding.delta = None
    finding.impacted = tuple(
        (ref, before[ref], after[ref]) for ref in watch if before[ref] != after[ref]
    )
    finding.sheet = sheet
    return finding


def main(path: str) -> int:
    print(f"Plumbline PoC  --  {Path(path).name}\n")
    sheets = load_formulas(path)
    total = 0

    for sheet, formulas in sheets.items():
        findings = detect_row_pattern_breaks(formulas)
        if not findings:
            continue
        watch = sorted(formulas)
        for f in findings:
            prove(path, sheet, f, watch)
            total += 1
            mark = "PROVED " if f.proved else "UNPROVED (dropped)"
            print(f"[{mark}] {sheet}!{f.cell}")
            print(f"    is        {f.actual}")
            print(f"    expected  {f.expected}")
            print(f"    why       {f.reason}")
            if f.proved:
                print(f"    proof     {f.cell}: {f.baseline_value} -> {f.repaired_value} "
                      f"(delta {f.delta:+})")
                for ref, b, a in f.impacted:
                    if ref == f.cell:
                        continue
                    print(f"              {ref}: {b} -> {a} (delta {a - b:+})")
            print()

    if total == 0:
        print("No pattern breaks found.")
    return 0


if __name__ == "__main__":
    # A stack trace is a fine answer for a library and a poor one for a demo script:
    # a reader who typos a path, or points this at a .xls, should get one line rather
    # than openpyxl internals.
    target = sys.argv[1] if len(sys.argv) > 1 else "tests/fixtures/quarterly_pl.xlsx"
    if not Path(target).exists():
        print("no such file: " + target, file=sys.stderr)
        raise SystemExit(2)
    try:
        raise SystemExit(main(target))
    except SystemExit:
        raise
    except Exception as exc:  # noqa: BLE001 -- one line, not a traceback
        print("could not read " + target + ": " + type(exc).__name__ + ": " + str(exc),
              file=sys.stderr)
        print("Plumbline reads .xlsx workbooks; older .xls files must be converted first.",
              file=sys.stderr)
        raise SystemExit(2) from None
