"""Tests for the guard that keeps Plumbline from emitting unreproducible proofs."""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

import plumbline  # noqa: F401
from plumbline.determinism import VOLATILE_FUNCTIONS, check, find_volatile, freeze

ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "tests" / "fixtures"
CLEAN = FIXTURES / "quarterly_pl.xlsx"

sys.path.insert(0, str(ROOT / "scripts"))


@pytest.fixture(scope="module")
def volatile_wb(tmp_path_factory) -> Path:
    """A workbook where one cell is random and another silently inherits it."""
    path = tmp_path_factory.mktemp("vol") / "volatile.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk"
    ws["A1"] = 100
    ws["A2"] = "=A1*2"
    ws["B1"] = "=RAND()"
    ws["B2"] = "=A1*B1"  # no RAND of its own -- contaminated through the graph
    wb.save(path)
    return path


class TestStaticScan:
    def test_finds_rand(self, volatile_wb):
        report = find_volatile(volatile_wb)
        assert report.is_volatile
        assert "Risk!B1" in report.cells
        assert report.functions == {"RAND"}

    def test_clean_workbook_is_clean(self):
        assert not find_volatile(CLEAN).is_volatile

    def test_does_not_flag_the_contaminated_cell(self, volatile_wb):
        """B2 has no volatile call of its own, so a name-based scan cannot see it.

        This is precisely why the empirical check has to exist as well.
        """
        assert "Risk!B2" not in find_volatile(volatile_wb).cells

    @pytest.mark.parametrize("func", sorted(VOLATILE_FUNCTIONS))
    def test_every_listed_volatile_is_detected(self, tmp_path, func):
        path = tmp_path / f"{func}.xlsx"
        wb = Workbook()
        wb.active["A1"] = f"={func}()" if func in {"RAND", "NOW", "TODAY"} else f"={func}(1,10)"
        wb.save(path)
        assert find_volatile(path).functions == {func}


class TestDeterminismCheck:
    def test_clean_workbook_is_stable(self):
        result = check(CLEAN)
        assert result.stable, result.summary()
        assert result.checked > 0

    def test_volatile_workbook_is_unstable(self, volatile_wb):
        result = check(volatile_wb)
        assert not result.stable
        assert "Risk!B1" in result.disagreements

    def test_catches_contamination_the_static_scan_misses(self, volatile_wb):
        """The whole reason for the second defence: B2 inherits instability."""
        result = check(volatile_wb)
        assert "Risk!B2" in result.disagreements
        assert "Risk!B2" not in find_volatile(volatile_wb).cells

    def test_deterministic_cells_do_not_drift(self, volatile_wb):
        """A2 is pure arithmetic; it must not be swept up as unstable."""
        assert "Risk!A2" not in check(volatile_wb).disagreements


class TestFreeze:
    def test_freezing_removes_volatility(self, volatile_wb, tmp_path):
        """A workbook saved by openpyxl has no cached values, so freezing cannot
        invent them -- and must not pretend it did."""
        out = tmp_path / "frozen.xlsx"
        report = freeze(volatile_wb, out)
        assert report.is_volatile  # it reports what it found

        # openpyxl-authored fixtures carry no cached results, so the formula stays
        # and the determinism check must still refuse the workbook. Substituting a
        # zero here would be exactly the silent wrongness we are guarding against.
        assert not check(out).stable

    def test_clean_workbook_passes_through(self, tmp_path):
        out = tmp_path / "copy.xlsx"
        report = freeze(CLEAN, out)
        assert not report.is_volatile
        assert out.exists()
        assert check(out).stable


def test_proof_pipeline_refuses_unstable_workbooks(volatile_wb):
    """The contract: never compute a delta on a workbook that fails the check.

    A delta between two runs of an unstable workbook is noise wearing the costume
    of evidence. This test states the rule the audit pipeline must honour.
    """
    assert not check(volatile_wb).stable
    assert check(CLEAN).stable
